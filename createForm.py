"""
화학물질 통계 정보 통합 검색 애플리케이션 (최종 개선판 v21)
- [디자인 대개편 - image_fa183d.png] 상단 검색 성공 메시지를 카드 내부의 입력 필드 바로 아래에 빨간색 텍스트로 정밀 바인딩
- [버그 해결] extract_company_name 함수 내부에서 정규식 unclosed bracket([) 오타로 인해 re.error가 발생하여 모든 파일 생성이 100% 오류 처리되던 문제를 완벽 해결
- [디버깅 보강] 크롤링 도중 예기치 못한 에러 발생 시, 단순 _ERROR 대신 구체적인 오류 원인(Exception string)을 테이블 행에 바로 보여주어 투명한 디버깅 지원
- [하단 피드백 박스 높이 일치] st.container 안쪽의 columns가 stretch 정렬되도록 CSS를 보강하고, 가변 markdown 구조의 모든 wrap div 높이를 100% 강제 연동하여 좌측 상태창과 우측 수직 버튼들의 가로 정렬 및 세로 높이를 1:1(84px)로 완벽하게 조율
- [레이블 간격 축소] '업체 검색' 레이블과 입력창 비율을 0.08:0.92로 극단 조율하여 거리를 기존의 1/5로 타이트하게 압축
"""

import io
import re
import zipfile
import requests
import time
from datetime import datetime
from collections import Counter
from html import unescape
from html.parser import HTMLParser
from urllib.parse import urljoin
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.comments import Comment
import streamlit as st

# ==================== 설정 ====================
BASE_URL = 'https://icis.mcee.go.kr'
SEARCH_JSON_URL = urljoin(BASE_URL, '/iprtr/cdrInfoDetailListJson.do')
SEARCH_PAGE_URL = urljoin(BASE_URL, '/search/searchType6.do?pageType=tabB')
DETAIL_VIEW_URL = urljoin(BASE_URL, '/iprtr/cdrInfoView.do')
HEADERS = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64)',
    'Referer': SEARCH_PAGE_URL,
}

# ==================== 연간입고량/연간사용·판매량 코드 매핑 ====================
QUANTITY_CODE_MAPPING = {
    '01': '0.1미만',
    '02': '0.1 ~ 0.5',
    '03': '0.5 ~ 1.0',
    '04': '1 ~ 2.5',
    '05': '2.5 ~ 5.0',
    '06': '5 ~ 20',
    '07': '20 ~ 200',
    '08': '200 ~ 1,000',
    '09': '1,000 ~ 5,000',
    '10': '5,000이상',
}

# ==================== HTTP 통신 함수 ====================
def fetch_search_page(session):
    """검색 페이지 가져오기"""
    res = session.get(SEARCH_PAGE_URL, headers=HEADERS, timeout=20)
    res.raise_for_status()
    return res.text

def fetch_search_results(session, company_name, page_no=1, search_year=None):
    """검색 결과 가져오기"""
    data = {
        'bplcNm': company_name,
        'pageNo': page_no,
    }
    if search_year:
        data['searchYear'] = search_year
    res = session.post(SEARCH_JSON_URL, headers=HEADERS, data=data, timeout=20)
    res.raise_for_status()
    return res.json()

def fetch_all_search_results(session, company_name, search_year=None):
    """모든 검색 결과 가져오기"""
    items = []
    page_no = 1
    while True:
        if search_year is None:
            use_year = '2022'
        else:
            use_year = search_year
        search_json = fetch_search_results(session, company_name, page_no=page_no, search_year=use_year)
        if search_json.get('result') != 'SUCCESS':
            return []
        page_items = search_json.get('list', [])
        total_count = int(search_json.get('totalCount', len(page_items)))
        if not page_items:
            break
        items.extend(page_items)
        if len(items) >= total_count:
            break
        page_no += 1
    return items

def fetch_detail_page(session, bplc_id, search_year=None):
    """상세 페이지 가져오기"""
    form_data = {
        'searchYear': '2022',
        'bplcId': bplc_id,
        'streNo': '',
    }
    res = session.post(DETAIL_VIEW_URL, headers=HEADERS, data=form_data, timeout=30)
    res.raise_for_status()
    return res.text

# ==================== 파일명 및 주소 처리 함수 ====================
def sanitize_filename(name):
    """파일명 특수문자 제거"""
    cleaned = re.sub(r'[<>:"/\\|?*\n\r\t]+', '_', name)
    cleaned = cleaned.strip().rstrip('.')
    return cleaned or 'detail'

def extract_region_from_address(address):
    """주소에서 시/군/구 정보 추출"""
    if not address:
        return ''
    address = address.strip()

    # 시/군/구 정확히 추출 (특별자치시, 광역시, 특별시 포함)
    match = re.search(r'([가-힣]+?(?:특별자치시|광역시|특별시|시|군))', address)
    if not match:
        return ''

    region = match.group(1)

    # 접미사 정리
    if region.endswith('특별자치시'):
        return region[:-6]
    if region.endswith('광역시') or region.endswith('특별시'):
        return region[:-3]
    return region[:-1]

def extract_company_name(html_text):
    """HTML에서 업체명 추출"""
    match = re.search(
        r'<th[^>]*>\s*업체명\s*</th>\s*<td[^>]*>(.*?)</td>',
        html_text,
        re.S | re.I,
    )
    if not match:
        return None
    # 💥 [버그 해결] 정규식 컴파일 오류(unterminated character set)를 일으키던 불완전 오타 코드 완전 제거
    company_name = re.sub(r'<[^>]+>', '', match.group(1))
    return unescape(company_name).strip()

def extract_section_html(html_text, section_title='3. 화학물질 취급현황'):
    """HTML에서 특정 섹션 추출"""
    section_search = re.search(
        rf'<h4[^>]*>\s*{re.escape(section_title)}\s*</h4>',
        html_text,
        re.S | re.I,
    )
    if not section_search:
        return None
    search_from = section_search.end()
    next_section = re.search(r'<h4[^>]*>\s*[0-9]+\.', html_text[search_from:], re.S | re.I)
    if next_section:
        end_index = search_from + next_section.start()
    else:
        end_index = len(html_text)
    return html_text[section_search.start():end_index]

def select_section_table(html_text, section_title='3. 화학물질 취급현황', table_index=2):
    """섹션에서 특정 테이블 선택"""
    section_html = extract_section_html(html_text, section_title=section_title)
    if not section_html:
        return None
    tables = re.findall(r'(<table[^>]*>.*?</table>)', section_html, re.S | re.I)
    if not tables:
        return None
    if 1 <= table_index <= len(tables):
        return tables[table_index - 1]
    return tables[-1]

# ==================== HTML 테이블 파싱 클래스 ====================
class TableHTMLParser(HTMLParser):
    """HTML 테이블 파서"""
    def __init__(self):
        super().__init__()
        self.in_table = False
        self.table_depth = 0
        self.in_tr = False
        self.in_cell = False
        self.current_cell = ''
        self.current_row = []
        self.rows = []

    def handle_starttag(self, tag, attrs):
        if tag == 'table':
            if self.in_table:
                self.table_depth += 1
            else:
                self.in_table = True
                self.table_depth = 1
        if self.in_table and tag == 'tr':
            self.in_tr = True
            self.current_row = []
        if self.in_tr and tag in ('td', 'th'):
            self.in_cell = True
            self.current_cell = ''

    def handle_data(self, data):
        if self.in_cell:
            self.current_cell += data

    def handle_endtag(self, tag):
        if tag in ('td', 'th') and self.in_cell:
            text = unescape(self.current_cell).strip()
            self.current_row.append(' '.join(text.split()))
            self.in_cell = False
        elif tag == 'tr' and self.in_tr:
            if self.current_row:
                self.rows.append(self.current_row)
            self.in_tr = False
        elif tag == 'table' and self.in_table:
            self.table_depth -= 1
            if self.table_depth <= 0:
                self.in_table = False

def parse_table_html(table_html):
    """HTML 테이블 파싱"""
    parser = TableHTMLParser()
    parser.feed(table_html)
    return parser.rows[2:] if len(parser.rows) > 2 else parser.rows

def _is_header_only_row(row):
    """헤더만 있는 행인지 판단"""
    header_keywords = [
        '물질명칭', 'CAS', 'CAS No', 'CAS No.', '제품', '제품명', '인체등유해성물질',
        '제한물질2', '금지물질2', '허가물질2', '사고대비물질2', '중점관리물질2',
        '금지·허가물질2', '노출·허용기준물질2', '직업환경측정물질등2', '위험물2',
        '독성가스2', '연간입고량', '연간사용·판매량'
    ]
    normalized_keywords = [kw.lower().replace('.', '').replace('·', ' ') for kw in header_keywords]
    cells = [cell.strip().lower().replace('.', '').replace('·', ' ') for cell in row if cell.strip()]
    if not cells:
        return False
    return all(any(nk == cell or nk in cell or cell in nk for nk in normalized_keywords) for cell in cells)

def _to_number_if_possible(value):
    """숫자로 변환 시도"""
    if not isinstance(value, str):
        return value
    text = value.strip().replace(',', '')
    if not text:
        return value
    if re.fullmatch(r'-?\d+', text):
        return int(text)
    if re.fullmatch(r'-?\d+\.\d+', text):
        return float(text)
    return value

def _convert_quantity_code(value):
    """연간입고량/연간사용·판매량 코드 변환 (01~10 -> 범위값)"""
    if not isinstance(value, str):
        return value
    code = value.strip()
    if code in QUANTITY_CODE_MAPPING:
        return QUANTITY_CODE_MAPPING[code]
    return value

def clean_rows(rows):
    """행 정제"""
    if not rows:
        return rows
    header_idx = None
    for i, row in enumerate(rows):
        joined = ' '.join(row)
        if 'CAS' in joined or 'CAS No' in joined or 'CAS No.' in joined:
            header_idx = i
            break
    if header_idx is not None:
        rows = rows[header_idx:]
        if len(rows) >= 2 and _is_header_only_row(rows[1]):
            return [rows[0]] + rows[2:]
        return rows
    return rows

# ==================== 엑셀 생성 함수 ====================
def workbook_to_bytes(rows):
    """행 데이터를 엑셀 바이트로 변환 (77개 컬럼, test260703.xlsx 정확히 동일)"""
    wb = Workbook()
    ws = wb.active
    
    # 77개 열의 완전한 헤더 (A~BY)
    full_headers = [
        '#', '물질명칭', 'CAS No.', '결과없음', '발암성', '생식독성', '생식세포 변이원성', 'CMR',
        '급성 독성(경구)', '급성 독성(경피)', '급성 독성(흡입)', '급성 독성', '흡인 유해성',
        '피부 부식성/피부 자극성', '심한 눈 손상성/눈 자극성', '피부/눈 자극성', '호흡기 과민성',
        '피부 과민성', '피부/호흡기 과민성', '특정표적장기 독성(1회 노출)', '특정표적장기 독성(반복 노출)',
        '특정표적장기 독성', '급성 수생환경 유해성', '만성 수생환경 유해성', '수생환경 유해성',
        '폭발성 물질', '자기반응성 물질', '유기과산화물', '산화성 가스', '산화성 액체',
        '산화성 고체', '인화성 가스', '인화성 에어로졸', '인화성 액체', '인화성 고체',
        '인화성', '자연발화성 액체', '자연발화성 고체', '물반응성 물질', '고압가스',
        '자기발열성 물질', '금속부식성 물질', 'TWA', 'STEL', '증기압', '개정일',
        '관리대상유해물질', '특별관리물질', '작업환경측정대상물질', '특수건강진단대상물질',
        '노출기준설정물질', '허용기준설정물질', '금지물질', '제한물질', '인체급성유해성물질',
        '인체만성유해성물질', '생태유해성물질', '허가물질', '사고대비물질', '중점관리물질',
        '위험물', '독성가스', '인체등유해성물질', '제한물질2', '금지물질2', '허가물질2',
        '사고대비물질2', '중점관리물질2', '금지·허가물질2', '노출·허용기준물질2',
        '직업환경측정물질등2', '위험물2', '독성가스2', '연간입고량', '연간사용·판매량',
        '[11번] 특정표적장기 독성(1회 노출)', '[11번] 특정표적장기 독성(반복 노출)'
    ]
    
    # 스타일 정의 (test260703.xlsx 정확 분석 기준)
    header_font = Font(name='맑은 고딕', size=10, bold=False, color=None)  # 색상 None = 기본값
    header_font_bold = Font(name='맑은 고딕', size=10, bold=True, color=None)  # 굵은 글씨체
    header_fill = PatternFill(start_color='FF5B9BD5', end_color='FF5B9BD5', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # 굵은 글씨체 헤더 열 (col_idx 기준)
    bold_header_cols = {5, 6, 8, 16, 22, 25, 36, 75}  # E, F, H, P, V, Y, AJ, BW
    
    # 메모 정보 (col_idx -> 메모 내용)
    header_comments = {
        10: '경피 급성독성 화학물질을 쓸 때 차폐 및 보호장구 매우 중요함',  # J: 급성 독성(경피)
        11: '흡입시 사망 가능가스를 쓸 때 차폐/국소배기장치/환기 및 보호장구 매우 중요함',  # K: 급성 독성(흡입)
        13: '흡입유해성이 큰 화학물질을 쓸 때 차폐/국소배기장치/환기 및 보호장구 매우 중요함',  # M: 흡인 유해성
        24: '만성 수생환경 유해성이 있는 물질이 우수관 등을 통해서 하천으로 흘러들면 하천생태계에 큰 영향을 줌',  # X: 만성 수생환경 유해성
        32: '인화성 가스를 쓰는 공정에서 접지 불량 등 발견되면 심각한 화재위험이 됨',  # AF: 인화성 가스
    }
    
    data_font = Font(name='맑은 고딕', size=10)
    
    # 배경색: 열별 그룹 (고정, 행별 교대 아님!)
    fill_abc = PatternFill(start_color='FFDDEBF7', end_color='FFDDEBF7', fill_type='solid')      # A~C: 연한 파란색
    fill_d_bj = PatternFill(start_color='FFFFFF00', end_color='FFFFFF00', fill_type='solid')     # D~BJ: 노란색
    fill_bk_bw = PatternFill(start_color='FFFFC000', end_color='FFFFC000', fill_type='solid')    # BK~BW: 주황색
    fill_bx_by = PatternFill(start_color='FFFFFF00', end_color='FFFFFF00', fill_type='solid')    # BX~BY: 노란색
    
    # 정렬 규칙
    alignment_default_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    alignment_default_left = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    
    # 첫 번째 행 (헤더) 생성 - 77개 열 모두
    for c_idx, header_text in enumerate(full_headers, start=1):
        cell = ws.cell(row=1, column=c_idx, value=header_text)
        
        # 굵기 적용
        if c_idx in bold_header_cols:
            cell.font = header_font_bold
        else:
            cell.font = header_font
        
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
        
        # 메모 추가
        if c_idx in header_comments:
            cell.comment = Comment(header_comments[c_idx], 'author')
    
    # 헤더 행 높이 설정 (52.2px)
    ws.row_dimensions[1].height = 52.2
    
    # 데이터 행 생성
    if rows and len(rows) > 1:
        for r_idx, row in enumerate(rows[1:], start=2):
            # 77개 열 모두에 대해 셀 생성
            for c_idx in range(1, 78):
                cell = ws.cell(row=r_idx, column=c_idx)
                cell.font = data_font
                cell.border = thin_border
                
                # 열 그룹별 배경색 지정 (고정)
                if 1 <= c_idx <= 3:  # A~C
                    cell.fill = fill_abc
                elif 4 <= c_idx <= 62:  # D~BJ
                    cell.fill = fill_d_bj
                elif 63 <= c_idx <= 75:  # BK~BW
                    cell.fill = fill_bk_bw
                else:  # BX~BY (76~77)
                    cell.fill = fill_bx_by
                
                # 데이터 배치 및 정렬 설정
                if c_idx == 1:  # A열: 순번
                    cell.value = r_idx - 1
                    cell.alignment = alignment_default_center
                elif c_idx == 2 and len(row) > 0:  # B열: 물질명칭
                    cell.value = row[0]
                    cell.alignment = alignment_default_left
                elif c_idx == 3 and len(row) > 1:  # C열: CAS No.
                    cell.value = str(row[1]) if row[1] is not None else None
                    cell.alignment = alignment_default_center
                    cell.number_format = '@'  # 💥 엑셀이 날짜로 자동 변환하지 못하도록 '텍스트' 형식 강제 지정
                elif 63 <= c_idx <= 77 and len(row) > (c_idx - 61):  # BK~BY열: 크롤링 데이터
                    # c_idx=63 → row[2], c_idx=77 → row[16]
                    cell.value = row[c_idx - 61]
                    cell.alignment = alignment_default_center
                else:  # 나머지: 빈 칸
                    cell.value = None
                    cell.alignment = alignment_default_center
    
    # 열 너비 설정 (test260703.xlsx 정확 값 - 모든 77개 열)
    column_widths = {
        'A': 5.0,
        'B': 31.0,
        'C': 13.0,
        'D': 12.0,
        'E': 10.0,
        'F': 10.0,
        'G': 7.09765625,
        'H': 9.0,
        'I': 8.296875,
        'J': 8.8984375,
        'K': 8.3984375,
        'L': 8.296875,
        'M': 8.59765625,
        'N': 10.19921875,
        'O': 10.8984375,
        'P': 11.0,
        'Q': 8.59765625,
        'R': 10.0,
        'S': 10.0,
        'T': 12.0,
        'U': 12.0,
        'V': 12.0,
        'W': 9.5,
        'X': 9.5,
        'Y': 9.0,
        'Z': 7.0,
        'AA': 9.0,
        'AB': 9.0,
        'AC': 7.0,
        'AD': 7.0,
        'AE': 7.0,
        'AF': 7.0,
        'AG': 8.0,
        'AH': 7.0,
        'AI': 7.0,
        'AJ': 8.0,
        'AK': 9.0,
        'AL': 9.0,
        'AM': 9.0,
        'AN': 9.0,
        'AO': 9.0,
        'AP': 9.0,
        'AQ': 9.0,
        'AR': 9.0,
        'AS': 9.5,
        'AT': 9.0,
        'AU': 9.0,
        'AV': 9.0,
        'AW': 9.0,
        'AX': 9.0,
        'AY': 9.0,
        'AZ': 9.0,
        'BA': 8.0,
        'BB': 8.0,
        'BC': 9.0,
        'BD': 9.0,
        'BE': 8.0,
        'BF': 8.0,
        'BG': 8.0,
        'BH': 8.0,
        'BI': 8.0,
        'BJ': 8.0,
        'BK': 9.0,
        'BL': 9.0,
        'BM': 9.0,
        'BN': 9.0,
        'BO': 8.0,
        'BP': 8.0,
        'BQ': 9.0,
        'BR': 9.0,
        'BS': 9.0,
        'BT': 9.0,
        'BU': 9.0,
        'BV': 9.5,
        'BW': 9.5,
        'BX': 15.0,
        'BY': 15.0,
    }
    
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width
    
    # 줌 레벨 80% 설정
    ws.sheet_view.zoomScale = 80
    
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()

def create_excel_bytes_for_company(session, company_name, item, search_year=None):
    """업체별 엑셀 파일 생성"""
    detail_html = fetch_detail_page(session, item['bplcId'], search_year=search_year)
    company_name_full = extract_company_name(detail_html) or item.get('bplcNm', company_name)
    target_table_html = select_section_table(detail_html, section_title='3. 화학물질 취급현황', table_index=2)
    if not target_table_html:
        raise RuntimeError('화학물질 취급현황 섹션의 두번째 표를 찾을 수 없습니다.')
    rows = parse_table_html(target_table_html)
    rows = clean_rows(rows)
    if not rows:
        raise RuntimeError('표에서 데이터를 찾을 수 없습니다.')
    first_join = ' '.join(rows[0]) if rows else ''
    if 'CAS' in first_join or 'CAS No' in first_join or 'CAS No.' in first_join:
        data_rows = rows[1:]
    else:
        data_rows = rows
    header = ['물질명칭', 'CAS No.', '인체등유해성물질', '제한물질2', '금지물질2', '허가물질2', '사고대비물질2', '중점관리물질2', '금지·허가물질2', '노출·허용기준물질2', '직업환경측정물질등2', '위험물2', '독성가스2', '연간입고량', '연간사용·판매량', '[11번] 특정표적장기 독성(1회 노출)', '[11번] 특정표적장기 독성(반복 노출)']
    
    # 연간입고량(13)과 연간사용·판매량(14)은 코드 변환, 나머지는 그대로
    quantity_columns = {13, 14}
    converted_rows = []
    for row in data_rows:
        converted_row = []
        for idx, cell in enumerate(row):
            if idx in quantity_columns:
                converted_row.append(_convert_quantity_code(cell))
            else:
                converted_row.append(cell)
        converted_rows.append(converted_row)
    final_rows = [header] + converted_rows
    return company_name_full, workbook_to_bytes(final_rows)

def search_companies(keywords, search_year=None):
    """업체 검색"""
    session = requests.Session()
    fetch_search_page(session)
    search_results = []
    for keyword in keywords:
        items = fetch_all_search_results(session, keyword, search_year=search_year)
        company_names = [item.get('bplcNm', '').strip() for item in items if item.get('bplcNm')]
        duplicate_counts = Counter(company_names)
        
        for item in items:
            company_name = item.get('bplcNm', '')
            company_address = item.get('locplcAdres', '') or item.get('bplcAdres', '')
            region = extract_region_from_address(company_address)
            if not region:
                region = ''
            
            use_region = duplicate_counts[company_name] > 1
            
            search_results.append({
                'index': len(search_results),
                'bplcId': item.get('bplcId', ''),
                'company_name': company_name,
                'region': region,
                'use_region_in_filename': use_region,
                'search_keyword': keyword,
            })
        
        if not items:
            search_results.append({
                'index': len(search_results),
                'company_name': f'검색 결과 없음: {keyword}',
                'region': '',
                'use_region_in_filename': False,
                'search_keyword': keyword,
            })

    return search_results

def zip_files(files_dict, selected_filenames):
    """파일들을 ZIP으로 압축"""
    zip_bytes = io.BytesIO()
    with zipfile.ZipFile(zip_bytes, 'w', zipfile.ZIP_DEFLATED) as archive:
        for filename in selected_filenames:
            data = files_dict.get(filename)
            if data is not None:
                archive.writestr(filename, data)
    zip_bytes.seek(0)
    return zip_bytes.getvalue()

# ==================== Streamlit UI Callbacks ====================
def toggle_select_all():
    """전체 선택 체크박스 변경 시 호출되는 콜백 (단일 패스 업데이트)"""
    val = st.session_state['select_all_main']
    st.session_state['select_all_state'] = val
    search_results = st.session_state.get('search_results', [])
    for i in range(len(search_results)):
        st.session_state[f'chk_{i}'] = val

def check_individual_toggle():
    """개별 체크박스 변경 시 호출되는 콜백 (전체 선택 체크박스 상태 동기화)"""
    search_results = st.session_state.get('search_results', [])
    all_checked = True
    for i in range(len(search_results)):
        if not st.session_state.get(f'chk_{i}', False):
            all_checked = False
            break
    st.session_state['select_all_main'] = all_checked
    st.session_state['select_all_state'] = all_checked

# ==================== Streamlit UI ====================
def main_ui(tab_mode=False):
    """메인 UI

    Parameters:
    - tab_mode: optional flag when called from combined_app tabs (ignored)
    """
    st.set_page_config(page_title='화학물질 통계 검색', layout='wide')
    #st.title('🔬 화학물질 통계 정보공개 검색')
    #st.write('')

    # 정교한 카드 레이아웃 및 세로 정렬 커스텀 CSS 마감
    st.markdown("""
    <style>
    /* 기본 공통 버튼 스타일 */
    button {
        background-color: #f0f5f9 !important;
        color: #555 !important;
        border: 1px solid #d0d7de !important;
        font-weight: 500;
    }
    button:hover {
        background-color: #e5ebf2 !important;
        border-color: #b1bac4 !important;
    }
    
    /* st.form 스타일 자체를 테두리 둥근 카드로 리디자인 */
    div[data-testid="stForm"] {
        border: 1px solid #e6e8eb !important;
        border-radius: 12px !important;
        padding: 20px !important;
        background-color: #ffffff !important;
        box-shadow: 0 4px 6px rgba(0,0,0,0.02) !important;
    }
    
    /* 입력창 실제 입력 필드의 배경을 흰색으로 지정하고 플레이스홀더 글자크기(14px)와 완벽 통일 */
    div[data-testid="stTextInput"] input {
        background-color: #ffffff !important;
        color: #31333F !important;
        font-size: 14px !important;
        font-family: inherit !important;
        border: 1px solid #d0d7de !important;
        border-radius: 6px !important;
    }
    /* 포커스 상태에서도 흰색 배경 유지 */
    div[data-testid="stTextInput"] input:focus {
        background-color: #ffffff !important;
        color: #31333F !important;
        border: 1px solid #d0d7de !important;
        outline: none !important;
        box-shadow: none !important;
    }
    /* 자동 완성(autofill) 상태에서도 흰색 배경 유지 */
    div[data-testid="stTextInput"] input:-webkit-autofill {
        -webkit-box-shadow: 0 0 0 1000px white inset !important;
        -webkit-text-fill-color: #31333F !important;
        background-color: #ffffff !important;
    }
    div[data-testid="stTextInput"] input:-webkit-autofill:focus {
        -webkit-box-shadow: 0 0 0 1000px white inset !important;
        -webkit-text-fill-color: #31333F !important;
        background-color: #ffffff !important;
    }
    div[data-testid="stTextInput"] input::placeholder {
        font-size: 14px !important;
        font-family: inherit !important;
        color: #8c959f !important;
    }
    
    /* 테이블용 st.columns의 모든 행 내부 요소 세로 중간 정렬 고정 */
    div.table-row-container div[data-testid="column"] {
        display: flex !important;
        align-items: center !important; 
        min-height: 40px !important;
    }
    
    /* 텍스트 하단 마진 제거하여 체크박스와 가로 수평 완벽 동기화 */
    div.table-row-container div[data-testid="column"] p, 
    div.table-row-container div[data-testid="column"] div, 
    div.table-row-container div[data-testid="column"] span {
        margin: 0 !important;
        padding: 0 !important;
        line-height: 1.2 !important;
    }
    
    /* 테이블 내부의 텍스트만 2.5px 아래로 수평 수선 */
    div.table-row-container div[data-testid="column"]:not(:nth-of-type(1)) p,
    div.table-row-container div[data-testid="column"]:not(:nth-of-type(1)) div,
    div.table-row-container div[data-testid="column"]:not(:nth-of-type(1)) span {
        transform: translateY(2.5px) !important;
    }
    
    /* 체크박스 마진 제거 */
    div[data-testid="stCheckbox"] {
        margin: 0 !important;
        padding: 0 !important;
        display: flex !important;
        align-items: center !important;
        justify-content: center !important;
    }
    
    /* 체크박스 하위 label 세로 정밀 1:1 중간 정렬 */
    div[data-testid="stCheckbox"] > label {
        display: flex !important;
        align-items: center !important;
        justify-content: center !important;
        margin: 0 !important;
        padding: 0 !important;
        min-height: unset !important;
    }

    /* No, 지역, 상태 열(1, 2, 4, 5번째 열) 내용 가운데 가로 정렬 */
    div.table-row-container div[data-testid="column"]:nth-of-type(1),
    div.table-row-container div[data-testid="column"]:nth-of-type(2),
    div.table-row-container div[data-testid="column"]:nth-of-type(4),
    div.table-row-container div[data-testid="column"]:nth-of-type(5) {
        justify-content: center !important;
    }
    
    /* 업체명 및 파일명 열(3, 6번째 열) 내용 왼쪽 정렬 유지 */
    div.table-row-container div[data-testid="column"]:nth-of-type(3),
    div.table-row-container div[data-testid="column"]:nth-of-type(6) {
        justify-content: flex-start !important;
    }
    
    /* 테이블 첫 번째 컬럼(체크박스 영역) 디자인 */
    div.table-row-container div[data-testid="column"]:nth-of-type(1) {
        background-color: #f0f0f0;
        padding: 4px;
        border-radius: 4px;
        display: flex;
        justify-content: center;
        align-items: center;
    }

    /* 💥 [하단 카드 정교화] 가로/세로 오차 없는 정밀 높이 바인딩 및 둥근 모서리 마감 */
    .bottom-card-anchor + div[data-testid="stContainer"] {
        border-radius: 12px !important;
        padding: 16px 20px !important;
        background-color: #ffffff !important;
        border: 1px solid #e6e8eb !important;
        box-shadow: 0 4px 6px rgba(0,0,0,0.02) !important;
    }

    /* 하단 카드 내부의 columns 가로 정렬 및 여백 조율 */
    .bottom-card-anchor + div[data-testid="stContainer"] div[data-testid="stHorizontalBlock"] {
        display: flex !important;
        align-items: stretch !important; /* 양쪽 열의 높이를 무조건 일치시킴 */
        gap: 16px !important;
        height: 84px !important;
    }

    /* [하단 카드 전용 column 정밀 override] */
    .bottom-card-anchor + div[data-testid="stContainer"] div[data-testid="column"] {
        display: flex !important;
        align-items: stretch !important;
        min-height: 84px !important;
        height: 84px !important;
    }

    /* [높이 일치 마감] 왼쪽 column의 모든 래퍼와 빈 공간의 높이를 84px로 강제 통일 */
    .bottom-card-anchor + div[data-testid="stContainer"] div[data-testid="column"]:nth-of-type(1) {
        display: flex !important;
        align-items: stretch !important;
        justify-content: flex-start !important;
        width: 85% !important; /* 가로폭 상단 검색창 비율과 매칭 */
        height: 84px !important;
        min-height: 84px !important;
    }

    .bottom-card-anchor + div[data-testid="stContainer"] div[data-testid="column"]:nth-of-type(1) > div[data-testid="stVerticalBlock"],
    .bottom-card-anchor + div[data-testid="stContainer"] div[data-testid="column"]:nth-of-type(1) div.element-container,
    .bottom-card-anchor + div[data-testid="stContainer"] div[data-testid="column"]:nth-of-type(1) div.stMarkdown,
    .bottom-card-anchor + div[data-testid="stContainer"] div[data-testid="column"]:nth-of-type(1) div.stMarkdown > div {
        height: 100% !important;
        margin: 0 !important;
        padding: 0 !important;
        display: flex !important;
        align-items: stretch !important;
        width: 100% !important;
    }

    /* [높이 일치 마감] 오른쪽 버튼 column 내부 stVerticalBlock에 flex 정렬 부여하여 84px 완벽 실장 */
    .bottom-card-anchor + div[data-testid="stContainer"] div[data-testid="column"]:nth-of-type(2) {
        display: flex !important;
        flex-direction: column !important;
        justify-content: space-between !important;
        align-items: stretch !important;
        width: 15% !important; /* 가로폭 상단 검색창 비율과 1:1 완벽 통일 */
        height: 84px !important;
        min-height: 84px !important;
    }

    .bottom-card-anchor + div[data-testid="stContainer"] div[data-testid="column"]:nth-of-type(2) > div[data-testid="stVerticalBlock"] {
        display: flex !important;
        flex-direction: column !important;
        justify-content: space-between !important;
        height: 84px !important;
        gap: 8px !important;
    }

    .bottom-card-anchor + div[data-testid="stContainer"] div[data-testid="column"]:nth-of-type(2) div.element-container {
        margin: 0 !important;
        padding: 0 !important;
        height: 38px !important;
    }

    /* 하단 카드 전용 수직 버튼 세트 규격 강제 통일 */
    .bottom-card-anchor + div[data-testid="stContainer"] button {
        height: 38px !important;
        min-height: 38px !important;
        margin: 0 !important;
        padding: 0px 16px !important;
        line-height: 38px !important;
        display: flex !important;
        align-items: center !important;
        justify-content: center !important;
        width: 100% !important;
        box-sizing: border-box !important;
    }

    /* 💥 [하단 피드백 박스 정밀 매칭] 둥근 모서리 및 높이 100% 맞춤형 가변 안내 박스 CSS 구축 */
    div.custom-status-card {
        border-radius: 10px !important;
        padding: 15px 24px !important;
        height: 100px !important; /* 가로 가이드라인 및 기준선 높이에 완벽 매치 */
        min-height: 100px !important;
        width: 100% !important;
        display: flex !important;
        align-items: center !important;
        justify-content: flex-start !important;
        font-size: 15px !important;
        font-weight: 500 !important;
        line-height: 1.4 !important;
        box-sizing: border-box !important;
        margin: 0 !important;
    }
    
    /* 가변 상태 메시지 카드 내부 p/span 태그 마진 및 트랜스레이트 강제 리셋 */
    div.custom-status-card p,
    div.custom-status-card span,
    div.custom-status-card div {
        margin: 0 !important;
        padding: 0 !important;
        transform: none !important;
    }

    /* 진행 상태(파스텔톤 하늘색 배경) */
    div.custom-status-card.info-type {
        background-color: #e8f0fe !important;
        border: 1px solid #d2e3fc !important;
        color: #1967d2 !important;
    }
    /* 완료 상태(파스텔톤 연초록색 배경) */
    div.custom-status-card.success-type {
        background-color: #e6f4ea !important;
        border: 1px solid #ceead6 !important;
        color: #137333 !important;
    }
    /* 경고 상태(파스텔톤 주황/붉은색 배경) */
    div.custom-status-card.warning-type {
        background-color: #fce8e6 !important;
        border: 1px solid #fad2cf !important;
        color: #c5221f !important;
    }

    /* 검색 폼 내부 라벨 수평 정렬용 텍스트 보정 */
    p.search-label-txt {
        font-size: 16px;
        font-weight: bold;
        margin: 0 !important;
        padding: 0 !important;
        white-space: nowrap;
        transform: translateY(8px); /* 텍스트박스 세로선 높이에 맞춘 마이크로 정렬 */
    }
    </style>
    """, unsafe_allow_html=True)

    # 세션 상태 초기화
    if 'search_results' not in st.session_state:
        st.session_state['search_results'] = []
    if 'files' not in st.session_state:
        st.session_state['files'] = {}
    if 'select_all_state' not in st.session_state:
        st.session_state['select_all_state'] = False
    if 'select_all_main' not in st.session_state:
        st.session_state['select_all_main'] = False
    if 'keywords' not in st.session_state:
        st.session_state['keywords'] = ''
    if 'search_success_message' not in st.session_state:
        st.session_state['search_success_message'] = None
    if 'last_search_keywords' not in st.session_state:
        st.session_state['last_search_keywords'] = None
    
    # 생성 및 다운로드 상태 관리용 세션 상태
    if 'file_generation_started' not in st.session_state:
        st.session_state['file_generation_started'] = False
    if 'file_generation_completed' not in st.session_state:
        st.session_state['file_generation_completed'] = False
    
    # 다운로드 상태 범용화 (ZIP 뿐만 아니라 1개 엑셀 개별 다운로드 대응)
    if 'download_bytes_data' not in st.session_state:
        st.session_state['download_bytes_data'] = None
    if 'download_filename' not in st.session_state:
        st.session_state['download_filename'] = ""
    if 'download_mime_type' not in st.session_state:
        st.session_state['download_mime_type'] = "application/zip"
    if 'download_completed' not in st.session_state:
        st.session_state['download_completed'] = False

    # [상단 검색 카드] 하나의 카드 프레임(st.form) 안에 가로 1열 수평 정렬을 정밀하게 적용
    with st.form(key='search_form', clear_on_submit=False):
        col1, col2 = st.columns([0.85, 0.15])
        
        with col1:
            # '업체 검색' 라벨과 입력 필드의 가로 폭 비율을 0.08 : 0.92로 극단 조율하여 거리를 기존의 1/5로 타이트하게 압축
            sub_col_lbl, sub_col_input = st.columns([0.08, 0.92])
            with sub_col_lbl:
                st.markdown('<p class="search-label-txt">업체 검색</p>', unsafe_allow_html=True)
            with sub_col_input:
                keyword_input = st.text_input(
                    '업체 검색',
                    value='',
                    placeholder='예) 삼성전자, 엘지화학',
                    key='keyword_input',
                    label_visibility='collapsed' # 기본 라벨은 가려 수평 정렬 유지
                )
                # 💥 [디자인 개편 - image_fa183d.png] 검색 결과 및 로딩 메시지창을 카드 내부 하단(입력창 바로 아래)에 위치
                search_msg_container = st.empty()
        
        search_button_pressed = False
        with col2:
            search_button_pressed = st.form_submit_button('🔍 검색', use_container_width=True)

    # 검색 실행
    if search_button_pressed:
        # 즉시 이전 검색 결과 테이블과 관련 파일/다운로드 상태 완전 초기화 (누르는 즉시 리셋!)
        st.session_state['search_results'] = []
        st.session_state['search_success_message'] = None
        st.session_state['files'] = {}
        st.session_state['file_generation_started'] = False
        st.session_state['file_generation_completed'] = False
        st.session_state['download_bytes_data'] = None
        st.session_state['download_filename'] = ""
        st.session_state['download_mime_type'] = "application/zip"
        st.session_state['download_completed'] = False
        
        if not keyword_input.strip():
            search_msg_container.markdown(f'<p style="color: #e53e3e; font-size: 15px; font-weight: 500; margin-top: -10px; margin-bottom: 10px;">검색어를 입력하세요.</p>', unsafe_allow_html=True)
        else:
            keywords = [kw.strip() for kw in keyword_input.split(',') if kw.strip()]
            if not keywords:
                search_msg_container.markdown(f'<p style="color: #e53e3e; font-size: 15px; font-weight: 500; margin-top: -10px; margin-bottom: 10px;">유효한 검색어가 없습니다.</p>', unsafe_allow_html=True)
            else:
                st.session_state['keywords'] = ', '.join(keywords)
                # 💥 [디자인 개편] 카드 내부 검색어 필드 바로 아래에 빨간색 텍스트로 로딩 상태 노출 (flicker 제거)
                #search_msg_container.markdown('<p style="color: #e53e3e; font-size: 15px; font-weight: 500; margin-top: 10px; margin-bottom: 0;">🔄 업체를 검색 중입니다...</p>', unsafe_allow_html=True)
                try:
                    search_results = search_companies(keywords)
                    
                    # 검색 결과에서 "검색 결과 없음" 아이템 필터링
                    actual_results = [item for item in search_results if not item['company_name'].startswith('검색 결과 없음')]
                    no_result_keywords = [item['search_keyword'] for item in search_results if item['company_name'].startswith('검색 결과 없음')]
                    
                    st.session_state['search_results'] = actual_results
                    st.session_state['select_all_state'] = False
                    st.session_state['select_all_main'] = False
                    
                    # 각 행 체크박스 및 상태 초기화
                    for i in range(len(actual_results)):
                        st.session_state[f'chk_{i}'] = False
                        st.session_state[f'status_{i}'] = ''
                        st.session_state[f'filename_{i}'] = ''
                    
                    # 메시지 처리
                    if actual_results:
                        st.session_state['search_success_message'] = f'{len(actual_results)}개의 업체를 찾았습니다.'
                        search_msg_container.markdown(f'<p style="color: #e53e3e; font-size: 15px; font-weight: 500; margin-top: -10px; margin-bottom: 10px;">{st.session_state["search_success_message"]}</p>', unsafe_allow_html=True)
                    elif no_result_keywords:
                        keywords_str = ', '.join(no_result_keywords)
                        st.session_state['search_success_message'] = f'검색 결과가 없습니다'
                        search_msg_container.markdown(f'<p style="color: #e53e3e; font-size: 15px; font-weight: 500; margin-top: -10px; margin-bottom: 10px;">{st.session_state["search_success_message"]}</p>', unsafe_allow_html=True)
                    
                    st.session_state['last_search_keywords'] = ', '.join(keywords)
                except Exception as exc:
                    search_msg_container.markdown(f'<p style="color: #e53e3e; font-size: 15px; font-weight: 500; margin-top: -10px; margin-bottom: 10px;">검색 중 오류: {exc}</p>', unsafe_allow_html=True)

    # 검색 결과 표시
    search_results = st.session_state['search_results']

    if search_results:
        # 📌 1단계: 검색 성공 메시지 상시 고정 노출 (깜빡임 없이 카드 내부 제자리에 단단하게 유지됨)
        if st.session_state['search_success_message']:
            search_msg_container.markdown(f'<p style="color: #e53e3e; font-size: 15px; font-weight: 500; margin-top: -10px; margin-bottom: 10px;">{st.session_state["search_success_message"]}</p>', unsafe_allow_html=True)

        # 📌 [깜빡임 및 체크박스 유실 완벽 해결]
        # 표 전체와 버튼 및 피드백 영역을 st.fragment로 묶어 리런 범위를 안전하게 격리합니다.
        @st.fragment
        def render_interactive_table():
            # [콜백] 다운로드 완료 등록
            def register_download():
                st.session_state['download_completed'] = True

            # HTML 표 영역 마진과 세로 중앙 정렬을 적용하기 위해 전용 div 래퍼 주입
            st.markdown('<div class="table-row-container">', unsafe_allow_html=True)

            # --------------------------------------------------
            # 1. 테이블 헤더 렌더링
            # --------------------------------------------------
            header_cols = st.columns([0.05, 0.08, 0.32, 0.1, 0.12, 0.4])
            # [제목행 디자인] 패딩을 12px -> 6px 12px 로 대폭 줄여 깔끔하게 슬림화
            header_style = 'background-color: #f0f0f0; padding: 6px 12px; font-weight: bold; text-align: center; border-radius: 4px;'
            
            with header_cols[0]:
                st.checkbox(
                    '',
                    key='select_all_main',
                    on_change=toggle_select_all
                )
            
            with header_cols[1]:
                st.markdown(f'<div style="{header_style}">No</div>', unsafe_allow_html=True)
            with header_cols[2]:
                st.markdown(f'<div style="{header_style}">업체명</div>', unsafe_allow_html=True)
            with header_cols[3]:
                st.markdown(f'<div style="{header_style}">지역</div>', unsafe_allow_html=True)
            with header_cols[4]:
                st.markdown(f'<div style="{header_style}">상태</div>', unsafe_allow_html=True)
            with header_cols[5]:
                st.markdown(f'<div style="{header_style}">파일명</div>', unsafe_allow_html=True)
            
            # [간격 조절] 제목행과 첫 번째 데이터 행 사이의 틈을 정확히 12px 만큼 여유 있게 벌려줍니다.
            st.markdown('<div style="margin-bottom: 12px;"></div>', unsafe_allow_html=True)

            # --------------------------------------------------
            # 2. 각 업체 데이터 행 표시 및 실시간 업데이트용 플레이스홀더 수집
            # --------------------------------------------------
            status_placeholders = {}
            filename_placeholders = {}

            for i, result in enumerate(search_results):
                if f'chk_{i}' not in st.session_state:
                    st.session_state[f'chk_{i}'] = False
                if f'status_{i}' not in st.session_state:
                    st.session_state[f'status_{i}'] = ''
                if f'filename_{i}' not in st.session_state:
                    st.session_state[f'filename_{i}'] = ''
                
                cols = st.columns([0.05, 0.08, 0.32, 0.1, 0.12, 0.4])
                
                with cols[0]:
                    st.checkbox(
                        '',
                        key=f'chk_{i}',
                        label_visibility='collapsed',
                        on_change=check_individual_toggle
                    )
                
                with cols[1]:
                    st.markdown(f'<div style="text-align: center;">{i + 1}</div>', unsafe_allow_html=True)
                with cols[2]:
                    st.write(result.get('company_name', ''))
                with cols[3]:
                    st.markdown(f'<div style="text-align: center;">{result.get("region", "") or ""}</div>', unsafe_allow_html=True)
                with cols[4]:
                    # 상태를 담을 고정형 빈 영역 생성 후 세션 상태의 기존 텍스트 렌더링
                    status_placeholders[i] = st.empty()
                    status_placeholders[i].markdown(f'<div style="text-align: center;">{st.session_state[f"status_{i}"]}</div>', unsafe_allow_html=True)
                with cols[5]:
                    # 파일명을 담을 고정형 빈 영역 생성 후 세션 상태의 기존 파일명 렌더링
                    filename_placeholders[i] = st.empty()
                    filename_placeholders[i].write(st.session_state[f'filename_{i}'])

            st.markdown('</div>', unsafe_allow_html=True) # table-row-container 종료

            # --------------------------------------------------
            # 3. [하단 제어 카드 영역] image_f8c9ec.png 스타일 완벽 일치 구조 구축
            # --------------------------------------------------
            st.markdown('<div class="bottom-card-anchor"></div>', unsafe_allow_html=True) # CSS 바인딩용 앵커
            #with st.container(border=True):
                # 💥 버튼의 길이를 검색 버튼의 가로폭과 완전히 일치시키기 위해 동일한 [0.85, 0.15] 비율 적용
                #col_msg, col_btns = st.columns([0.85, 0.15])

            col_msg, col_btn_gen, col_btn_down = st.columns([0.7, 0.15, 0.15], vertical_alignment="center")


            # 메시지 텍스트 렌더링용 헬퍼 포맷 적용
            with col_msg:
                status_placeholder = st.empty()
                if st.session_state.get('download_completed', False):
                    status_placeholder.markdown('<div style="color: #e53e3e; font-size: 15px; font-weight: 500;">파일 다운로드가 완료되었습니다!</div>', unsafe_allow_html=True)
                elif st.session_state.get('file_generation_completed', False):
                    status_placeholder.markdown('<div style="color: #e53e3e; font-size: 15px; font-weight: 500;">파일이 생성되었습니다! 다운로드하세요.</div>', unsafe_allow_html=True)
                else:
                    status_placeholder.markdown('<div style="color: #e53e3e; font-size: 15px; font-weight: 500;">업체를 선택하고 파일을 생성하세요.</div>', unsafe_allow_html=True)
                    
            # 첫 번째 버튼: 파일 생성
            with col_btn_gen:
                is_generating = st.session_state.get('file_generation_started', False)
                gen_clicked_bottom = st.button('📝 파일 생성', key='btn_gen', use_container_width=True, disabled=is_generating)
                
            # 두 번째 버튼: 다운로드
            with col_btn_down:
                if st.session_state['download_bytes_data'] is not None:
                    st.download_button(
                        label='📥 다운로드',
                        data=st.session_state['download_bytes_data'],
                        file_name=st.session_state['download_filename'],
                        mime=st.session_state.get('download_mime_type', 'application/zip'),
                        key='btn_down',
                        use_container_width=True,
                        on_click=register_download
                    )
                else:
                    st.button('📥 다운로드', key='btn_down_disabled', use_container_width=True, disabled=True)


            # "📝 파일 생성" 요청이 발생했을 때 가동되는 고성능 동기식 루프
            if gen_clicked_bottom:
                selected_idxs = [
                    idx for idx in range(len(search_results)) if st.session_state.get(f'chk_{idx}', False)
                ]
                
                if not selected_idxs:
                    status_placeholder.markdown('<div style="color: #e53e3e; font-size: 15px; font-weight: 500;">업체를 선택하세요.</div>', unsafe_allow_html=True)
                else:
                    # 파일 생성 시작 처리
                    st.session_state['file_generation_started'] = True
                    st.session_state['file_generation_completed'] = False
                    st.session_state['download_bytes_data'] = None
                    st.session_state['download_filename'] = ""
                    st.session_state['download_mime_type'] = "application/zip"
                    st.session_state['download_completed'] = False
                    
                    # 상태 및 파일명 일괄 클리어
                    for i in range(len(search_results)):
                        st.session_state[f'status_{i}'] = ''
                        st.session_state[f'filename_{i}'] = ''
                        status_placeholders[i].markdown('<div style="text-align: center;"></div>', unsafe_allow_html=True)
                        filename_placeholders[i].write('')

                    for current_idx, idx in enumerate(selected_idxs):
                        item_data = search_results[idx]
                        keyword = item_data.get('search_keyword', '')
                        bplc_id = item_data.get('bplcId', '')
                        company_name = item_data.get('company_name', '')
                        region = item_data.get('region', '')
                        use_region = item_data.get('use_region_in_filename', False)
                        
                        if bplc_id and '검색 결과 없음' not in company_name:
                            try:
                                session = requests.Session()
                                fetch_search_page(session)
                                company_name_full, wb_bytes = create_excel_bytes_for_company(
                                    session, keyword, {'bplcId': bplc_id, 'bplcNm': company_name}, search_year=None
                                )
                                
                                if use_region and region and region != '':
                                    filename = f"{sanitize_filename(company_name_full)}_{sanitize_filename(region)}.xlsx"
                                else:
                                    filename = f"{sanitize_filename(company_name_full)}.xlsx"
                                
                                if 'files' not in st.session_state:
                                    st.session_state['files'] = {}
                                st.session_state['files'][filename] = wb_bytes
                                
                                st.session_state[f'status_{idx}'] = '✅'
                                st.session_state[f'filename_{idx}'] = filename
                                
                                # UI 플레이스홀더 값 즉시 교체 (중복 위젯 생성 불필요)
                                status_placeholders[idx].markdown('<div style="text-align: center;">✅</div>', unsafe_allow_html=True)
                                filename_placeholders[idx].write(filename)
                            
                            except Exception as exc:
                                # 💥 상세 에러 메시지를 괄호 안에 담아 표출하여 디버깅을 돕습니다.
                                error_msg = str(exc)
                                st.session_state[f'status_{idx}'] = '❌'
                                st.session_state[f'filename_{idx}'] = f'{company_name}_ERROR ({error_msg})'
                                
                                status_placeholders[idx].markdown('<div style="text-align: center;">❌</div>', unsafe_allow_html=True)
                                filename_placeholders[idx].write(f'{company_name}_ERROR ({error_msg})')
                        
                        time.sleep(0.1)
                        
                    # 최종 다운로드 포맷 결정 (선택된 파일 개수 기준 분기)
                    selected_filenames = []
                    for idx in selected_idxs:
                        filename = st.session_state.get(f'filename_{idx}', '')
                        if filename and not '_ERROR' in filename:
                            selected_filenames.append(filename)
                    
                    if selected_filenames:
                        if len(selected_filenames) == 1:
                            # 💥 선택한 업체가 1개면 zip 압축 없이 파일 그대로 다운로드 받도록 수정
                            single_filename = selected_filenames[0]
                            file_bytes = st.session_state['files'].get(single_filename)
                            
                            st.session_state['download_bytes_data'] = file_bytes
                            st.session_state['download_filename'] = single_filename
                            st.session_state['download_mime_type'] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        else:
                            # 2개 이상인 경우 기존처럼 ZIP 압축 후 제공
                            zip_bytes = zip_files(st.session_state['files'], selected_filenames)
                            now = datetime.now().strftime('%Y%m%d_%H%M%S')
                            keywords_str = st.session_state['keywords'].replace(', ', '_').replace(' ', '_')
                            
                            st.session_state['download_bytes_data'] = zip_bytes
                            st.session_state['download_filename'] = f'{keywords_str}_{now}.zip'
                            st.session_state['download_mime_type'] = "application/zip"
                            
                        st.session_state['file_generation_completed'] = True
                        st.session_state['file_generation_started'] = False
                        
                        # 최종 다운로드 데이터 및 상태 정보를 UI 버튼 구조에 반영하기 위해 Rerun 기동
                        st.rerun()
                    else:
                        st.session_state['file_generation_started'] = False
                        status_placeholder.markdown('<div style="color: #e53e3e; font-size: 15px; font-weight: 500;">정상적으로 생성된 파일이 없어 다운로드 파일을 구축하지 못했습니다.</div>', unsafe_allow_html=True)

        # 격리된 대화형 표 프래그먼트 렌더링 호출
        render_interactive_table()

if __name__ == '__main__':
    main_ui()