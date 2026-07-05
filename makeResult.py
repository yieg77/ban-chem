# ----------------- 라이브러리 임포트 -----------------
import streamlit as st
import pandas as pd
import requests
import xml.etree.ElementTree as ET
from collections import defaultdict
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from io import BytesIO
import re
import zipfile
import base64
from openpyxl.styles import PatternFill  #20260420
from openpyxl.styles.colors import Color #20260420
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side, Font, Alignment
import concurrent.futures  # 멀티스레딩을 위한 라이브러리 추가
import html  # html.unescape 사용을 위해 전역 임포트

# ----------------- 고정 유해성 항목 순서 -----------------
HAZARD_ORDER = [
    '#', '물질명칭', 'CAS No.', '결과없음',
    '발암성', '생식독성', '생식세포 변이원성', 'CMR',
    '급성 독성(경구)', '급성 독성(경피)', '급성 독성(흡입)', '급성 독성',  #260419
    '흡인 유해성', '피부 부식성/피부 자극성', '심한 눈 손상성/눈 자극성', '피부/눈 자극성', #260419
    '호흡기 과민성', '피부 과민성', '피부/호흡기 과민성', #260419
    '특정표적장기 독성(1회 노출)', '특정표적장기 독성(반복 노출)', '특정표적장기 독성', #260419
    '급성 수생환경 유해성', '만성 수생환경 유해성', '수생환경 유해성', #260419
    '폭발성 물질', '자기반응성 물질', '유기과산화물', '산화성 가스',
    '산화성 액체', '산화성 고체', '인화성 가스', '인화성 에어로졸',
    '인화성 액체', '인화성 고체', 
    '인화성',  #260630
    '자연발화성 액체', '자연발화성 고체',
    '물반응성 물질', '고압가스', '자기발열성 물질', '금속부식성 물질',
    'TWA', 'STEL', '증기압', '개정일',
    '관리대상유해물질', '특별관리물질', 
    '작업환경측정대상물질', #0930
    '특수건강진단대상물질',
    '노출기준설정물질','허용기준설정물질','금지물질','제한물질', # 260630
    '인체급성유해성물질', '인체만성유해성물질', '생태유해성물질',# 260630
    '허가물질','사고대비물질','중점관리물질','위험물','독성가스', #0930
    '[11번] 특정표적장기 독성(1회 노출)', '[11번] 특정표적장기 독성(반복 노출)'  #260419
]

# 260419 ########################
# 전역상수로 변경
# -------------------- 등급 판별 기준 --------------------
CMR_GRADE_PRIORITY = {'1A': 0, '1B': 1, '2': 2}
GRADE_PRIORITY = {'1': 0, '2': 1, '3': 2, '4': 3}
#260701 폰트 정의
default_font = Font(name='맑은 고딕', size=10, bold=False)
bold_font    = Font(name='맑은 고딕', size=10, bold=True)
#------------------------260701

# ----------------- 취합 항목 정의 -----------------
AGGREGATE_GROUPS = [
    ('급성 독성',      ['급성 독성(경구)', '급성 독성(경피)', '급성 독성(흡입)']),
    ('피부/눈 자극성', ['피부 부식성/피부 자극성', '심한 눈 손상성/눈 자극성']),
    ('피부/호흡기 과민성', ['호흡기 과민성', '피부 과민성']),
    ('특정표적장기 독성',  ['특정표적장기 독성(1회 노출)', '특정표적장기 독성(반복 노출)']),
    ('수생환경 유해성',    ['급성 수생환경 유해성', '만성 수생환경 유해성']),
    ('인화성',    ['인화성 가스', '인화성 에어로졸', '인화성 액체', '인화성 고체']),  #260701
]

#260630 -------------------- 색상 정의 --------------------
fill_header = PatternFill(fill_type='solid', fgColor='FF5B9BD5')  # 제목행
fill_label  = PatternFill(fill_type='solid', fgColor='FFDDEBF7')  # 라벨열
fill_data1  = PatternFill(fill_type='solid', fgColor='FFFFFF00')  # 데이터셀1
fill_data2  = PatternFill(fill_type='solid', fgColor='FFFFC000')  # 데이터셀2
#------------------------260630

# ----------------- CMR 등급 추출 함수 -----------------
def extract_cmr_grades(val):
    if not isinstance(val, str) or not val.strip():
        return []
    grades = []
    for part in re.split(r'[\n|,]+', val):
        m = re.match(r'^\s*(1A|1B|2)\b', part.strip())
        if m:
            grades.append(m.group(1))
    return grades

# ----------------- 일반용 등급 추출 함수 -----------------
def extract_grades(val):
    if not isinstance(val, str) or not val.strip():
        return []
    grades = []
    for part in re.split(r'[\n|,]+', val):
        m = re.match(r'^\s*(1|2|3|4)\b', part.strip())
        if m:
            grades.append(m.group(1))
    return grades

def get_highest_cmr_grade(grades):
    valid = [g for g in grades if g in CMR_GRADE_PRIORITY]
    if not valid:
        return None
    return min(valid, key=lambda g: CMR_GRADE_PRIORITY[g])

def get_highest_grade(grades):
    valid = [g for g in grades if g in GRADE_PRIORITY]
    if not valid:
        return None
    return min(valid, key=lambda g: GRADE_PRIORITY[g])

def compute_cmr_grade(result, source_keys):
    grades = []
    for k in source_keys:
        grades.extend(extract_cmr_grades(result.get(k, '')))
    return get_highest_cmr_grade(grades)

def compute_aggregate_grade(result, source_keys):
    grades = []
    for k in source_keys:
        grades.extend(extract_grades(result.get(k, '')))
    return get_highest_grade(grades)

# ----------------- TWA, STEL 조회 함수 -----------------
# Session 객체를 인자로 받아 연결을 재사용하도록 개선
def query_twa_stel(session, service_key, chem_id):
    try:
        res = session.get(
            'https://msds.kosha.or.kr/openapi/service/msdschem/chemdetail08',
            params={'serviceKey': service_key, 'chemId': chem_id},
            timeout=10
        )
        res.encoding = 'utf-8'
        root = ET.fromstring(res.text)

        twa = ''
        stel = ''

        for item in root.findall('.//item'):
            name_twa = item.findtext('msdsItemNameKor')
            if name_twa == '국내규정':
                detail = item.findtext('itemDetail')
                if detail:
                    parts = [p.strip() for p in detail.split('|') if p.strip()]
                    for part in parts:
                        if part.startswith('TWA'):
                            match = re.search(r'TWA\s*[:]?\s*([\d\.]+\s*ppm(?:\([^)]*\))?)', part)
                            if match:
                                twa = match.group(1).strip()
                        elif part.startswith('STEL'):
                            match = re.search(r'STEL\s*[:]?\s*([\d\.]+\s*ppm(?:\([^)]*\))?)', part)
                            if match:
                                stel = match.group(1).strip()

        return twa, stel
    except:
        return '', ''

# ----------------- 증기압 조회 함수 -----------------
def query_vapor_pressure(session, service_key, chem_id):
    try:
        res = session.get(
            'https://msds.kosha.or.kr/openapi/service/msdschem/chemdetail09',
            params={'serviceKey': service_key, 'chemId': chem_id},
            timeout=10
        )
        res.encoding = 'utf-8'
        root = ET.fromstring(res.text)

        for item in root.findall('.//item'):
            name = item.findtext('msdsItemNameKor')
            if name and name.strip() == '증기압':
                detail = item.findtext('itemDetail')
                detail = html.unescape(detail)
                if detail and detail.strip():
                    cleaned = ''.join(detail.split())  # 공백 제거
                    cleaned = re.split(r'\|+|※+', cleaned)[0]
                    return cleaned
    except:
        pass
    return ''

# ----------------- 개정일 조회 함수 -----------------
def query_revision_date(session, service_key, chem_id):
    try:
        res = session.get(
            'https://msds.kosha.or.kr/openapi/service/msdschem/chemdetail16',
            params={'serviceKey': service_key, 'chemId': chem_id},
            timeout=10
        )
        res.encoding = 'utf-8'
        root = ET.fromstring(res.text)
        for item in root.findall('.//item'):
            if item.findtext('msdsItemNameKor') == '최종 개정일자':
                detail = item.findtext('itemDetail')
                return detail.strip() if detail else ''
    except:
        return ''
    return ''

# --- '관리대상유해물질', '특별관리물질' 등 조회 함수 ---
def has_keyword(detail, keyword):
    if not detail:
        return False
    items = [d.strip() for d in detail.split('|')]
    return any(keyword in item for item in items)

def query_detail15(session, service_key, chem_id):
    try:
        res = session.get(
            'https://msds.kosha.or.kr/openapi/service/msdschem/chemdetail15',
            params={'serviceKey': service_key, 'chemId': chem_id},
            timeout=10
        )
        res.encoding = 'utf-8'
        root = ET.fromstring(res.text)

        result = {
            '관리대상유해물질': False, '특별관리물질': False, '작업환경측정대상물질': False, '특수건강진단대상물질': False,
            '노출기준설정물질': False, '허용기준설정물질': False, '금지물질': False, '제한물질': False,
            '허가물질': False, '사고대비물질': False, '중점관리물질': False, '위험물': False, '독성가스': False,
            '인체급성유해성물질': False, '인체만성유해성물질': False, '생태유해성물질': False,
        }

        for item in root.findall('.//item'):
            code = item.findtext('msdsItemCode')
            detail = item.findtext('itemDetail') or ''

            if code in ['002', 'O04', 'O12', 'O06']:
                if has_keyword(detail, '금지물질'): result['금지물질'] = True
                if has_keyword(detail, '제한물질'): result['제한물질'] = True
                if has_keyword(detail, '허가물질'): result['허가물질'] = True
                if has_keyword(detail, '사고대비물질'): result['사고대비물질'] = True
                if has_keyword(detail, '중점관리물질'): result['중점관리물질'] = True
                if has_keyword(detail, '위험물'): result['위험물'] = True
                if has_keyword(detail, '독성가스'): result['독성가스'] = True

            if code == 'O02':
                if has_keyword(detail, '관리대상유해물질'): result['관리대상유해물질'] = True
                if has_keyword(detail, '특별관리물질'): result['특별관리물질'] = True
                if has_keyword(detail, '특수건강진단대상물질'): result['특수건강진단대상물질'] = True
                if has_keyword(detail, '작업환경측정대상물질'): result['작업환경측정대상물질'] = True
                if has_keyword(detail, '노출기준설정물질'): result['노출기준설정물질'] = True
                if has_keyword(detail, '허용기준설정물질'): result['허용기준설정물질'] = True

            elif code in ['O04', 'O12']:
                if has_keyword(detail, '인체급성유해성물질'): result['인체급성유해성물질'] = True
                if has_keyword(detail, '인체만성유해성물질'): result['인체만성유해성물질'] = True
                if has_keyword(detail, '생태유해성물질'): result['생태유해성물질'] = True
                    
        return result

    except Exception as e:
        return {
            '관리대상유해물질': False, '특별관리물질': False, '작업환경측정대상물질': False, '특수건강진단대상물질': False,
            '노출기준설정물질': False, '허용기준설정물질': False, '금지물질': False, '제한물질': False,
            '허가물질': False, '사고대비물질': False, '중점관리물질': False, '위험물': False, '독성가스': False,
            '인체급성유해성물질': False, '인체만성유해성물질': False, '생태유해성물질': False,
        }
        
def query_detail11(session, service_key, chem_id):
    try:
        res = session.get(
            'https://msds.kosha.or.kr/openapi/service/msdschem/chemdetail11',
            params={'serviceKey': service_key, 'chemId': chem_id},
            timeout=10
        )
        res.encoding = 'utf-8'
        root = ET.fromstring(res.text)

        stot_single   = ''
        stot_repeated = ''
        
        for item in root.findall('.//item'):
            code   = item.findtext('msdsItemCode')
            detail = html.unescape(item.findtext('itemDetail') or '')
            if detail.strip() == '자료없음':
                continue
            if code == 'K0418':
                stot_single = detail.strip()
            elif code == 'K0420':
                stot_repeated = detail.strip()

        return stot_single, stot_repeated
    except:
        return '', ''

# ----------------- 단일 화학물질 처리용 워커 함수 (멀티스레딩 지원) -----------------
def process_single_chemical(idx, row, session, service_key):
    # pandas 변환 과정에서 생성될 수 있는 'nan', 'None' 문자열 방어
    cas_raw = str(row[2]).strip()
    if cas_raw.lower() in ['nan', 'none', 'null']:
        cas = ''
    else:
        cas = cas_raw

    id_num = row[0]
    name = row[1]
    result = {'#': id_num, '물질명칭': name, 'CAS No.': cas}
    unknown_cols = set()
    
    # 💥 [CAS No. 예외 처리 추가] 💥
    if not cas:  # CAS No가 아예 비어있는 경우
        result['결과없음'] = '영업비밀'
        return idx, result, unknown_cols
    elif '심의중' in cas:  # 문자열 내에 '심의중'이 포함된 경우
        result['결과없음'] = '심의중'
        return idx, result, unknown_cols
    
    try:
        res_id = session.get(
            'https://msds.kosha.or.kr/openapi/service/msdschem/chemlist',
            params={'serviceKey': service_key, 'searchWrd': cas, 'searchCnd': 1},
            timeout=10
        )
        res_id.encoding = 'utf-8'
        chem_id = ET.fromstring(res_id.text).findtext('.//chemId')

        if not chem_id:
            result['결과없음'] = '공단 MSDS 없음'
        else:
            res_detail = session.get(
                'https://msds.kosha.or.kr/openapi/service/msdschem/chemdetail02',
                params={'serviceKey': service_key, 'chemId': chem_id},
                timeout=10
            )
            res_detail.encoding = 'utf-8'
            root = ET.fromstring(res_detail.text)

            b02_detail = next(
                (item.findtext('itemDetail')
                 for item in root.findall('.//item')
                 if item.findtext('msdsItemCode') == 'B02'),
                None
            )

            if b02_detail is None or b02_detail.strip() == '' or b02_detail.strip() == '자료없음':
                result['결과없음'] = '자료 없음'
            else:
                merged = defaultdict(list)
                inhalation_labels = [
                    '급성 독성(흡입)', '급성 독성(흡입: 가스)',
                    '급성 독성(흡입: 분진/미스트)', '급성 독성(흡입: 증기)'
                ]
                inhalation_entries = []
                cmr_map = {'발암성': [], '생식독성': [], '생식세포 변이원성': []}

                for entry in b02_detail.split('|'):
                    if ':' in entry and '자료없음' not in entry:
                        k, v = map(str.strip, entry.rsplit(':', 1))
                        v = v.replace('구분', '').strip()
                        if k in inhalation_labels:
                            label = k.replace('급성 독성(', '').replace(')', '')
                            inhalation_entries.append(f"{v}({label})")
                        elif k in cmr_map:
                            if not (cas == '64-17-5' and k == '발암성'):
                                cmr_map[k].append(v)
                                if v not in merged[k]:
                                    merged[k].append(v)
                        else:
                            if v not in merged[k]:
                                merged[k].append(v)

                if inhalation_entries:
                    result['급성 독성(흡입)'] = '\n'.join(inhalation_entries)

                for k, v_list in merged.items():
                    if cas == '64-17-5' and k == '발암성':
                        continue

                    result[k] = '\n'.join(v_list)
                    if k not in HAZARD_ORDER:
                        unknown_cols.add(k)

                cmr_grades = []
                for values in cmr_map.values():
                    for v in values:
                        cmr_grades.extend(extract_cmr_grades(v))

                highest_grade = get_highest_cmr_grade(cmr_grades)
                if highest_grade:
                    result['CMR'] = highest_grade

                for agg_col, source_keys in AGGREGATE_GROUPS:
                    grade = compute_aggregate_grade(result, source_keys)
                    if grade:
                        result[agg_col] = grade

            # 상세 정보 병합
            result['TWA'], result['STEL'] = query_twa_stel(session, service_key, chem_id)
            result['증기압'] = query_vapor_pressure(session, service_key, chem_id)
            result['개정일'] = query_revision_date(session, service_key, chem_id)

            res_detail15 = query_detail15(session, service_key, chem_id)
            if res_detail15['관리대상유해물질']: result['관리대상유해물질'] = '▣'
            if res_detail15['특별관리물질']: result['특별관리물질'] = '▣'
            if res_detail15['특수건강진단대상물질']: result['특수건강진단대상물질'] = '▣'
            if res_detail15['인체급성유해성물질']: result['인체급성유해성물질'] = '▣'
            if res_detail15['인체만성유해성물질']: result['인체만성유해성물질'] = '▣'
            if res_detail15['생태유해성물질']: result['생태유해성물질'] = '▣'
            if res_detail15['작업환경측정대상물질']: result['작업환경측정대상물질'] = '▣'
            if res_detail15['노출기준설정물질']: result['노출기준설정물질'] = '▣'
            if res_detail15['허용기준설정물질']: result['허용기준설정물질'] = '▣'
            if res_detail15['금지물질']: result['금지물질'] = '▣'
            if res_detail15['제한물질']: result['제한물질'] = '▣'
            if res_detail15['허가물질']: result['허가물질'] = '▣'
            if res_detail15['사고대비물질']: result['사고대비물질'] = '▣'
            if res_detail15['중점관리물질']: result['중점관리물질'] = '▣'
            if res_detail15['위험물']: result['위험물'] = '▣'
            if res_detail15['독성가스']: result['독성가스'] = '▣'
            
            stot_single, stot_repeated = query_detail11(session, service_key, chem_id)
            if stot_single: result['[11번] 특정표적장기 독성(1회 노출)'] = stot_single
            if stot_repeated: result['[11번] 특정표적장기 독성(반복 노출)'] = stot_repeated

    except Exception as e:
        result['결과없음'] = f'조회 오류: {str(e)}'

    return idx, result, unknown_cols

# ----------------- CAS 정보 조회 함수 (멀티스레딩 적용) -----------------
def query_cas_info(data_rows, service_key, progress_bar=None, progress_label=None, progress_text=None):
    results = []
    unknown_columns = set()
    total = len(data_rows)
    progress = progress_bar if progress_bar is not None else st.progress(0)
    
    processed_data = []
    completed = 0

    # 세션을 열고 재사용하여 연결 오버헤드 최소화
    with requests.Session() as session:
        # 최대 10개의 워커(스레드)를 띄워 병렬 통신 수행
        with concurrent.futures.ThreadPoolExecutor(max_workers=10) as executor:
            future_to_idx = {
                executor.submit(process_single_chemical, idx, row, session, service_key): idx
                for idx, row in enumerate(data_rows.itertuples(index=False), start=1)
            }
            
            # 완료되는 대로 결과 수집 및 프로그래스바 안전 업데이트
            for future in concurrent.futures.as_completed(future_to_idx):
                idx, res, unk_cols = future.result()
                processed_data.append((idx, res))
                unknown_columns.update(unk_cols)
                
                completed += 1
                ratio = min(completed / total, 1.0) if total else 1.0
                percent = int(ratio * 100)
                if progress_bar is not None:
                    progress.progress(ratio)
                    if progress_text is not None:
                        # [개선] 퍼센트 텍스트를 오른쪽 정렬하여 진행바 바로 왼쪽에 바짝 붙임
                        progress_text.markdown(
                            f"<div style='text-align: right;'>{percent}%</div>", 
                            unsafe_allow_html=True
                        )
                else:
                    progress.progress(ratio)
                    
    # 비동기 처리로 섞인 순서를 원래 인덱스 순서대로 재정렬
    processed_data.sort(key=lambda x: x[0])
    results = [x[1] for x in processed_data]

    df = pd.DataFrame(results)
    return df, sorted(list(unknown_columns))


# ----------------- Streamlit 앱 실행 -----------------
import os
import streamlit as st
SERVICE_KEY = st.secrets["SERVICE_KEY"]

def _render_download_button(label, data_bytes, file_name, mime):
    if not data_bytes:
        return
    st.download_button(
        label,
        data=data_bytes,
        file_name=file_name,
        mime=mime,
        key=f"download_{file_name}",
        use_container_width=False,
    )


def makeResult_ui():
    """유해성 정보 수집 UI - 단일 파일 처리"""
    if 'mr_processed' not in st.session_state:
        st.session_state.mr_processed = False
    if 'mr_result_file' not in st.session_state:
        st.session_state.mr_result_file = None
    if 'mr_uploader_key' not in st.session_state:
        st.session_state.mr_uploader_key = 0

    uploaded_file = st.file_uploader(
        "📎 엑셀 파일을 업로드 하세요! (입력파일명: A.xlsx, A: 회사명)",
        type="xlsx",
        key=f"mr_file_uploader_{st.session_state.mr_uploader_key}"
    )

    if uploaded_file and not st.session_state.mr_processed:
        _run_makeResult_body(uploaded_file)

    if st.session_state.mr_processed:
        st.success("✅ 유해성 정보 수집이 완료되었습니다.")
        
        input_filename = uploaded_file.name if uploaded_file else "output"
        basename = os.path.splitext(input_filename)[0]
        output_filename = f"{basename}_유해성분석.xlsx"
        
        # [개선] 버튼 너비가 글자 크기에 딱 맞도록 여유 공간(비율 2, 2, 6)을 확보하여 찌그러짐 방지
        col1, col2, col3 = st.columns([2, 2, 6])
        with col1:
            _render_download_button(
                "📥 결과 엑셀 다운로드",
                st.session_state.mr_result_file.getvalue() if st.session_state.mr_result_file is not None else None,
                output_filename,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        with col2:
            if st.button("🔁 새 파일 업로드", key="mr_reset"):
                st.session_state.mr_processed = False
                st.session_state.mr_result_file = None
                st.session_state.mr_uploader_key += 1
                st.rerun()


def makeResult_multi_ui():
    """유해성 정보 수집 UI - 다중 파일 처리"""
    
    # 세션 상태 초기화 (기존 로직 동일)
    if 'mr_files_list' not in st.session_state:
        st.session_state.mr_files_list = []
    if 'mr_processing_results' not in st.session_state:
        st.session_state.mr_processing_results = {}
    if 'mr_download_zip' not in st.session_state:
        st.session_state.mr_download_zip = None
    if 'mr_download_single' not in st.session_state:
        st.session_state.mr_download_single = None
    if 'mr_download_filename' not in st.session_state:
        st.session_state.mr_download_filename = None
    if 'mr_uploader_key' not in st.session_state:
        st.session_state.mr_uploader_key = 0
    if 'mr_processing_started' not in st.session_state:
        st.session_state.mr_processing_started = False
    if 'mr_upload_handled_key' not in st.session_state:
        st.session_state.mr_upload_handled_key = None
    if 'mr_uploaded_files' not in st.session_state:
        st.session_state.mr_uploaded_files = []

    uploader_placeholder = st.empty()

    # 1. 파일 업로드 영역
    with uploader_placeholder:
        uploaded_files = st.file_uploader(
            "엑셀 파일을 업로드하세요 (입력파일명: A.xlsx, A: 회사명)",
            type="xlsx",
            accept_multiple_files=True,
            key=f"mr_multi_uploader_{st.session_state.mr_uploader_key}",
        )

    # 2. 파일 업로드 이벤트 처리 (업로드 직후 업로더를 숨기고 처리 루프로 진입)
    if uploaded_files and st.session_state.mr_upload_handled_key != st.session_state.mr_uploader_key:
        st.session_state.mr_upload_handled_key = st.session_state.mr_uploader_key
        st.session_state.mr_processing_results = {}
        st.session_state.mr_download_zip = None
        st.session_state.mr_download_single = None
        st.session_state.mr_download_filename = None
        st.session_state.mr_processing_started = False
        
        st.session_state.mr_uploaded_files = [(f.name, f.read()) for f in uploaded_files]
        st.session_state.mr_files_list = [
            (name, BytesIO(data)) for name, data in st.session_state.mr_uploaded_files
        ]
        
        for fname, _ in st.session_state.mr_files_list:
            st.session_state.mr_processing_results[fname] = {
                'status': '⏳ 대기중', 'progress': 0, 'wb': None, 'excel_bytes': None
            }
            
        st.session_state.mr_uploader_key += 1
        uploader_placeholder.empty() # 업로드가 완료되면 업로더 UI를 깔끔하게 비움
        st.rerun()

    # 3. 파일 처리 현황 및 진행 상태 표시 영역
    if st.session_state.mr_files_list:
        st.markdown(
            f'<div style="color: #e53e3e; font-size: 15px; font-weight: 500; margin-top: -10px; margin-bottom: 15px;">'
            f'총 {len(st.session_state.mr_files_list)}개의 파일이 업로드되었습니다.'
            f'</div>', 
            unsafe_allow_html=True
        )
        st.markdown("### 📊 파일 처리 현황")

        progress_text_placeholders = {}
        progress_bar_placeholders = {}
        
        for fname, _ in st.session_state.mr_files_list:
            col1, col2, col3 = st.columns([2.4, 0.4, 4.2])
            with col1:
                st.markdown(f"<div>{fname}</div>", unsafe_allow_html=True)
            with col2:
                text_ph = st.empty()
                progress_text_placeholders[fname] = text_ph
            with col3:
                bar_ph = st.empty()
                progress_bar_placeholders[fname] = bar_ph

            result = st.session_state.mr_processing_results.get(fname, {})
            
            if result.get('status') == '✅ 완료':
                text_ph.markdown("<div style='text-align: right;'>100%</div>", unsafe_allow_html=True)
                bar_ph.progress(1.0)
            else:
                text_ph.markdown("<div style='text-align: right;'>0%</div>", unsafe_allow_html=True)
                bar_ph.progress(0.0)

        st.markdown("<br>", unsafe_allow_html=True)

        completed_count = sum(1 for r in st.session_state.mr_processing_results.values() if r.get('status', '').startswith('✅'))
        total_count = len(st.session_state.mr_files_list)
        st.progress(completed_count / total_count if total_count else 0.0, text=f"전체 진행 상태: {completed_count} / {total_count}")

        # 4. 순차적 파일 처리 로직 (이하 동일)
        if not st.session_state.mr_processing_started:
            st.session_state.mr_processing_started = True
            st.rerun()

        target = None
        for fname, file_obj in st.session_state.mr_files_list:
            result = st.session_state.mr_processing_results[fname]
            if not (result.get('status', '').startswith('✅') or result.get('status', '').startswith('❌')):
                target = (fname, file_obj)
                break

        if target:
            fname, file_obj = target
            result = st.session_state.mr_processing_results[fname]
            try:
                result['status'] = '처리중'
                result['progress'] = 0
                file_obj.seek(0)
                output, wb = _process_uploaded_file(
                    file_obj,
                    progress_bar=progress_bar_placeholders[fname],
                    progress_label=fname,
                    progress_text=progress_text_placeholders[fname],
                )
                result['status'] = '✅ 완료'
                result['wb'] = wb
                result['excel_bytes'] = output.getvalue()
                result['progress'] = 100
                st.rerun()
            except Exception as e:
                result['status'] = '❌ 오류'
                result['progress'] = 0
                st.rerun()

        else:
            _create_download_files()

        # 5. 다운로드 버튼 영역
        if st.session_state.mr_download_single is not None or st.session_state.mr_download_zip is not None:
            st.markdown("<br>", unsafe_allow_html=True)
            dl_col1, dl_col2, dl_col3 = st.columns([2.5, 2.5, 5])
            with dl_col1:
                if st.session_state.mr_download_single is not None:
                    _render_download_button(
                        "📥 파일 다운로드",
                        st.session_state.mr_download_single,
                        st.session_state.mr_download_filename,
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )
                elif st.session_state.mr_download_zip is not None:
                    _render_download_button(
                        "📥 파일 다운로드",
                        st.session_state.mr_download_zip,
                        f"화학물질_유해성_{pd.Timestamp.now().strftime('%Y%m%d_%H%M%S')}.zip",
                        "application/zip",
                    )

def _process_uploaded_file(uploaded_file, progress_bar=None, progress_label=None, progress_text=None):
    """업로드된 엑셀 파일을 처리해서 결과 워크북을 반환"""
    uploaded_file.seek(0)
    wb = load_workbook(uploaded_file)
    ws = wb.active
    uploaded_file.seek(0)
    raw_df = pd.read_excel(uploaded_file, header=None)
    header_row_full = raw_df.iloc[0].tolist()
    data_rows = raw_df[1:].copy()
    data_rows.columns = header_row_full
    header_row = header_row_full[:62]
    current_headers = set(header_row)
    SKIP_HEADER_CHECK = {'[11번] 특정표적장기 독성(1회 노출)', '[11번] 특정표적장기 독성(반복 노출)'}
    expected_headers = set(HAZARD_ORDER) - SKIP_HEADER_CHECK
    missing_headers = [h for h in HAZARD_ORDER if h not in current_headers and h not in SKIP_HEADER_CHECK]
    unexpected_headers = [h for h in header_row if h not in expected_headers]
    
    if unexpected_headers or missing_headers:
        st.error("❗제목행(A~AN열)에 오류가 있습니다. 유해성 정보를 조회하지 않습니다.")
        if unexpected_headers:
            st.markdown("### 🚫 예기치 않은 열 제목")
            for col in unexpected_headers:
                st.markdown(f"- `{col}`")
        if missing_headers:
            st.markdown("### ⚠️ 누락된 필수 항목")
            for col in missing_headers:
                st.markdown(f"- `{col}`")
        raise ValueError("헤더 검증 실패")

    required_cols = {'#', '물질명칭', 'CAS No.'}
    if not required_cols.issubset(set(data_rows.columns)):
        st.error("필수 열('#', '물질명칭', 'CAS No.')이 누락되어 있습니다.")
        raise ValueError("필수 열 누락")

    data_rows['CAS No.'] = data_rows['CAS No.'].astype(str).str.lstrip('0')
    data_rows = data_rows[['#', '물질명칭', 'CAS No.']].copy()
    hazard_df, _ = query_cas_info(
        data_rows,
        SERVICE_KEY,
        progress_bar=progress_bar,
        progress_label=progress_label,
        progress_text=progress_text,
    )
    _write_excel_results(wb, ws, hazard_df)
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output, wb


def _run_makeResult_body(uploaded_file):
    """유해성 정보 처리 본체 (makeResult_ui에서 호출)"""
    output, _ = _process_uploaded_file(uploaded_file)
    st.session_state.mr_result_file = output
    st.session_state.mr_processed = True


def _create_download_files():
    """완료된 파일을 엑셀/ZIP으로 묶어 다운로드 상태를 갱신"""
    completed_files = {
        fname: result for fname, result in st.session_state.mr_processing_results.items()
        if result.get('status') == '✅ 완료' and result.get('excel_bytes') is not None
    }

    if len(completed_files) == 1:
        fname, result = list(completed_files.items())[0]
        output_fname = fname.replace('.xlsx', '_유해성분석.xlsx')
        st.session_state.mr_download_single = result['excel_bytes']
        st.session_state.mr_download_filename = output_fname
        st.session_state.mr_download_zip = None
    elif len(completed_files) > 1:
        zip_buffer = BytesIO()
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zf:
            for fname, result in completed_files.items():
                output_fname = fname.replace('.xlsx', '_유해성분석.xlsx')
                zf.writestr(output_fname, result['excel_bytes'])
        zip_buffer.seek(0)
        st.session_state.mr_download_zip = zip_buffer.getvalue()
        st.session_state.mr_download_single = None
        st.session_state.mr_download_filename = None

def _write_excel_results(wb, ws, hazard_df):
    """처리 결과를 엑셀에 기록"""
    import math
    col_name_to_idx = {cell.value: idx for idx, cell in enumerate(ws[1], start=1)}
    for row in ws['A1:BY1']:
        for cell in row:
            cell.fill = fill_header
            cell.font = default_font
            if cell.value in ['발암성','생식독성','CMR','급성 독성','피부/눈 자극성','특정표적장기 독성','수생환경 유해성','인화성','연간사용·판매량']:
                cell.font = bold_font
    for r_idx, row in hazard_df.iterrows():
        excel_row = r_idx + 2
        for col_name in HAZARD_ORDER:
            if col_name in col_name_to_idx:
                col_idx = col_name_to_idx[col_name]
                cell = ws.cell(row=excel_row, column=col_idx)
                cell.value = row.get(col_name, '')
                col_letter = get_column_letter(col_idx)
                if col_letter in ['A', 'B', 'C']:
                    cell.fill = fill_label
                elif col_letter in ['BK', 'BL', 'BM', 'BN', 'BO', 'BP', 'BQ', 'BR', 'BS', 'BT', 'BU', 'BV','BW']:
                    cell.fill = fill_data2
                else:
                    cell.fill = fill_data1
                if col_name in ['#', 'CAS No.', '결과없음', '개정일',
                                '관리대상유해물질', '특별관리물질', '특수건강진단대상물질', '인체급성유해성물질', '인체만성유해성물질',
                                '생태유해성물질', '작업환경측정대상물질', '노출기준설정물질','허용기준설정물질','금지물질','제한물질',
                                '허가물질','사고대비물질','중점관리물질','위험물','독성가스']:
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                else:
                    if col_name in ['[11번] 특정표적장기 독성(1회 노출)', '[11번] 특정표적장기 독성(반복 노출)']:
                        cell.alignment = Alignment(vertical='top', wrap_text=False)
                    else:
                        cell.alignment = Alignment(vertical='center', wrap_text=True)
    _build_summary_tables(ws, hazard_df)

def _build_summary_tables(ws, hazard_df):
    """표2, 표3, 표4 생성"""
    import math
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    start_idx = HAZARD_ORDER.index('발암성')
    end_idx = HAZARD_ORDER.index('금속부식성 물질') + 1
    hazard_cols = HAZARD_ORDER[start_idx:end_idx]
    hazard_start_col = 4
    start_row = 2
    end_row = start_row + len(hazard_df) - 1
    summary_start_row = end_row + 2
    analyzed_count = sum(
        1 for r in range(start_row, end_row + 1)
        if ws.cell(row=r, column=HAZARD_ORDER.index('결과없음') + 1).value != '공단 MSDS 없음'
    )
    # 표2 제목행
    ws[f"D{summary_start_row}"] = "유해성"
    ws[f"D{summary_start_row}"].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws[f"D{summary_start_row}"].font = default_font
    ws[f"D{summary_start_row}"].border = thin_border
    ws[f"D{summary_start_row}"].fill = fill_header
    for idx, col_name in enumerate(hazard_cols):
        col_letter = get_column_letter(hazard_start_col + idx + 1)
        cell = ws[f"{col_letter}{summary_start_row}"]
        cell.value = col_name
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.font = bold_font if col_name in ['발암성','생식독성','CMR','급성 독성','피부/눈 자극성','특정표적장기 독성','수생환경 유해성','인화성'] else default_font
        cell.border = thin_border
        cell.fill = fill_header
    row_labels = ['구분1', '1A', '1B', '구분2', '구분3', '구분4', '기타구분', '유해물질수', '분석물질수', '유해물질비율']
    cmr_agg_cols = {'CMR'}
    general_agg_cols = {agg_col for agg_col, _ in AGGREGATE_GROUPS}
    cmr_source_cols = {'발암성', '생식독성', '생식세포 변이원성'}
    summary_data = []
    for hazard in hazard_cols:
        col_idx = HAZARD_ORDER.index(hazard) + 1
        count_map = {label: 0 for label in row_labels[:-3]}
        for r in range(start_row, end_row + 1):
            val = ws.cell(row=r, column=col_idx).value
            if val is None or (isinstance(val, float) and math.isnan(val)) or str(val).strip() == '':
                continue
            val_str = str(val).strip()
            if hazard in cmr_agg_cols:
                label_map = {'1A': '1A', '1B': '1B', '2': '구분2'}
                count_map[label_map.get(val_str, '기타구분')] += 1
            elif hazard in general_agg_cols:
                label_map = {'1': '구분1', '2': '구분2', '3': '구분3', '4': '구분4'}
                count_map[label_map.get(val_str, '기타구분')] += 1
            elif hazard in cmr_source_cols:
                grades = extract_cmr_grades(val_str)
                most_severe = get_highest_cmr_grade(grades)
                label_map = {'1A': '1A', '1B': '1B', '2': '구분2'}
                if most_severe:
                    count_map[label_map.get(most_severe, '기타구분')] += 1
            else:
                grades = extract_grades(val_str)
                most_severe = get_highest_grade(grades)
                label_map = {'1': '구분1', '2': '구분2', '3': '구분3', '4': '구분4'}
                if most_severe:
                    count_map[label_map.get(most_severe, '기타구분')] += 1
                else:
                    count_map['기타구분'] += 1
        count_map['유해물질수'] = sum(count_map[label] for label in row_labels[:7])
        count_map['분석물질수'] = analyzed_count
        count_map['유해물질비율'] = f"{round((count_map['유해물질수'] / analyzed_count) * 100)}%" if analyzed_count else "0%"
        for i, label in enumerate(row_labels):
            if len(summary_data) <= i:
                summary_data.append([])
            summary_data[i].append(count_map[label])
    for row_offset, (label, row_values) in enumerate(zip(row_labels, summary_data), start=1):
        row_num = summary_start_row + row_offset
        label_cell = ws.cell(row=row_num, column=hazard_start_col)
        label_cell.value = label
        label_cell.alignment = Alignment(horizontal='center', vertical='center')
        label_cell.font = default_font
        label_cell.border = thin_border
        label_cell.fill = fill_label
        for col_offset, value in enumerate(row_values):
            col_num = hazard_start_col + col_offset + 1
            cell = ws.cell(row=row_num, column=col_num)
            if label in ('유해물질수', '유해물질비율'):
                cell.value = value
            else:
                cell.value = value if value != 0 else None
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.font = default_font
            cell.border = thin_border
    # 표3
    summary_titles = [
        '관리대상유해물질', '특별관리물질', '작업환경측정대상물질', '특수건강진단대상물질',
        '노출기준설정물질','허용기준설정물질','금지물질','제한물질',
        '인체급성유해성물질', '인체만성유해성물질', '생태유해성물질',
        '허가물질','사고대비물질','중점관리물질','위험물','독성가스',
        '인체등유해성물질',
        '제한물질2', '금지물질2', '허가물질2', '사고대비물질2',
        '중점관리물질2', '금지·허가물질2', '노출·허용기준물질2', '직업환경측정물질등2', '위험물2', '독성가스2'
    ]
    summary_start_col = 47
    table3_start_row = end_row + 2
    analyzed_count2 = 0
    col_result_idx = HAZARD_ORDER.index('결과없음') + 1
    for r in range(start_row, end_row + 1):
        val = ws.cell(row=r, column=col_result_idx).value
        if val != '공단 MSDS 없음':
            analyzed_count2 += 1
    c = ws.cell(row=table3_start_row, column=summary_start_col - 1)
    c.value = "규제물질"; c.alignment = Alignment(horizontal='center', vertical='center')
    c.font = default_font; c.border = thin_border; c.fill = fill_header
    c = ws.cell(row=table3_start_row + 1, column=summary_start_col - 1)
    c.value = "물질 수"; c.alignment = Alignment(horizontal='center', vertical='center')
    c.font = default_font; c.border = thin_border; c.fill = fill_label
    c = ws.cell(row=table3_start_row + 2, column=summary_start_col - 1)
    c.value = "물질 비율"; c.alignment = Alignment(horizontal='center', vertical='center')
    c.font = default_font; c.border = thin_border; c.fill = fill_label
    for idx, col_name in enumerate(summary_titles):
        cell = ws.cell(row=table3_start_row, column=summary_start_col + idx, value=col_name)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.font = default_font; cell.border = thin_border; cell.fill = fill_header
    for idx in range(len(summary_titles)):
        col_idx = summary_start_col + idx
        count = sum(1 for r in range(start_row, end_row + 1) if str(ws.cell(row=r, column=col_idx).value).strip() == '▣')
        cell = ws.cell(row=table3_start_row + 1, column=col_idx, value=count)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.font = default_font; cell.border = thin_border
        ratio = f"{round((count / analyzed_count2) * 100)}%" if analyzed_count2 else "0%"
        cell = ws.cell(row=table3_start_row + 2, column=col_idx, value=ratio)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.font = default_font; cell.border = thin_border
    # 표4
    def _normalize_header(value):
        if value is None:
            return ''
        return str(value).strip().replace(' ', '')

    col_idx_in = col_idx_use = None
    for col in range(1, ws.max_column + 1):
        header = ws.cell(row=1, column=col).value
        norm = _normalize_header(header)
        if norm == '연간입고량':
            col_idx_in = col
        elif norm in {'연간사용·판매량', '연간사용판매량'}:
            col_idx_use = col
    if col_idx_in is None or col_idx_use is None:
        return
    table3_end_row = table3_start_row + 2
    table4_start_row = table3_end_row + 2
    table4_start_col = 73
    headers4 = ['중량(톤/년) 또는 부피단위(㎥/년)', '연간입고량', '연간사용·판매량']
    for idx, header in enumerate(headers4):
        cell = ws.cell(row=table4_start_row, column=table4_start_col + idx)
        cell.value = header; cell.font = default_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border; cell.fill = fill_header
    usage_map2 = {
        '1': "0.1미만", '2': "0.1~0.5", '3': "0.5~1.0", '4': "1~2.5", '5': "2.5~5.0",
        '6': "5~20", '7': "20~200", '8': "200~1,000", '9': "1,000~5,000", '10': "5,000이상",
        1: "0.1미만", 2: "0.1~0.5", 3: "0.5~1.0", 4: "1~2.5", 5: "2.5~5.0",
        6: "5~20", 7: "20~200", 8: "200~1,000", 9: "1,000~5,000", 10: "5,000이상",
        1.0: "0.1미만", 2.0: "0.1~0.5", 3.0: "0.5~1.0", 4.0: "1~2.5", 5.0: "2.5~5.0",
        6.0: "5~20", 7.0: "20~200", 8.0: "200~1,000", 9.0: "1,000~5,000", 10.0: "5,000이상",
    }
    usage_descriptions = ["0.1미만","0.1~0.5","0.5~1.0","1~2.5","2.5~5.0","5~20","20~200","200~1,000","1,000~5,000","5,000이상"]
    from collections import Counter as _Counter
    import math as _math
    def _norm(v):
        if v is None:
            return None
        if isinstance(v, bool):
            return None
        if isinstance(v, (int, float)) and not isinstance(v, bool):
            if v in usage_map2: return usage_map2[v]
            try: return usage_map2.get(int(v), None)
            except Exception: return None
        raw = str(v).strip()
        if raw == '': return None
        norm = raw.replace('\u3000', ' ').replace('，', ',').replace('–', '-').replace('\u2013', '-')
        norm = norm.replace(' ', '')
        for desc in usage_descriptions:
            if desc.replace(',', '') in norm or desc.replace('，','') in norm: return desc
        if '미만' in raw: return '0.1미만'
        if '이상' in raw: return '5,000이상'
        m = re.search(r'\b([1-9]|10)\b', raw)
        if m:
            try: return usage_map2.get(int(m.group(1)))
            except Exception: pass
        m2 = re.match(r'^\s*(\d+)(?:\.0+)?\s*$', raw)
        if m2:
            try: return usage_map2.get(int(m2.group(1)))
            except Exception: pass
        digits = re.sub(r'[^0-9]', '', raw)
        if digits:
            try:
                di = int(digits)
                if 1 <= di <= 10: return usage_map2.get(di)
            except Exception: pass
        return None
        
    incoming_counter = _Counter()
    usage_counter = _Counter()
    for r in range(start_row, end_row + 1):
        desc_in = _norm(ws.cell(row=r, column=col_idx_in).value)
        desc_use = _norm(ws.cell(row=r, column=col_idx_use).value)
        if desc_in:
            ws.cell(row=r, column=col_idx_in).value = desc_in
            incoming_counter[desc_in] += 1
        if desc_use:
            ws.cell(row=r, column=col_idx_use).value = desc_use
            usage_counter[desc_use] += 1
    for i, desc in enumerate(usage_descriptions):
        row = table4_start_row + 1 + i
        c_desc = ws.cell(row=row, column=table4_start_col)
        c_desc.value = desc; c_desc.font = default_font
        c_desc.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c_desc.border = thin_border; c_desc.fill = fill_label
        v_in = incoming_counter.get(desc, 0)
        c_in = ws.cell(row=row, column=table4_start_col + 1)
        c_in.value = v_in if v_in != 0 else None; c_in.font = default_font
        c_in.alignment = Alignment(horizontal='center', vertical='center'); c_in.border = thin_border
        v_use = usage_counter.get(desc, 0)
        c_use = ws.cell(row=row, column=table4_start_col + 2)
        c_use.value = v_use if v_use != 0 else None; c_use.font = default_font
        c_use.alignment = Alignment(horizontal='center', vertical='center'); c_use.border = thin_border

# standalone 실행용 (streamlit run makeResult.py) - import 시에는 실행 안됨
if __name__ != 'makeResult':
    makeResult_ui()
