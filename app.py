"""
화학물질 정보 수집 시스템 (단일 파일 버전)

요구사항 반영 요약
1) ICIS에서 업체별 취급현황 표를 가져와 기본 엑셀(77열)을 만든다.
2) 생성 직후 같은 메모리에서 바로 KOSHA MSDS API를 조회해 빈칸(D~BJ, BX~BY)을 채운다.
3) 같은 워크북에 표2~표4 요약표를 즉시 생성한다.
4) 사용자는 업체별 최종 파일을 한 번만 다운로드한다.

주의
- 기존 createForm/makeResult의 데이터 규칙을 유지하기 위해 핵심 로직을 그대로 재사용했다.
- 이해를 돕기 위해 주요 흐름에 상세 주석을 추가했다.

Streamlit 버전 : 1.58
"""

# ---------------------------
# 표준/외부 라이브러리 임포트
# ---------------------------
import io
import re
import time
import html
import random
import asyncio
import importlib
import zipfile
import requests
import pandas as pd
import streamlit as st
import concurrent.futures
import xml.etree.ElementTree as ET

try:
	aiohttp = importlib.import_module('aiohttp')
except Exception:  # aiohttp 미설치 환경에서는 기존 동기 경로를 사용
	aiohttp = None

from datetime import datetime
from collections import Counter, defaultdict
from html import unescape
from html.parser import HTMLParser
from urllib.parse import urljoin

from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ---------------------------
# 페이지 기본 설정
# ---------------------------
st.set_page_config(
	page_title='화학물질 정보 수집 시스템 (ICIS + KOSHA)',
	layout='wide',
	initial_sidebar_state='collapsed',
)


# ---------------------------
# 공통 상수 정의 (ICIS)
# ---------------------------
BASE_URL = 'https://icis.mcee.go.kr'
SEARCH_JSON_URL = urljoin(BASE_URL, '/iprtr/cdrInfoDetailListJson.do')
SEARCH_PAGE_URL = urljoin(BASE_URL, '/search/searchType6.do?pageType=tabB')
DETAIL_VIEW_URL = urljoin(BASE_URL, '/iprtr/cdrInfoView.do')
HEADERS = {
	'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64)',
	'Referer': SEARCH_PAGE_URL,
}

RETRY_STATUS_CODES = {429, 500, 502, 503, 504}
ASYNC_MAX_CONCURRENCY = 30


def request_with_retry(session, method, url, *, timeout=20, max_retries=3, backoff_base=0.6, retry_status_codes=None, **kwargs):
	"""일시적 네트워크 오류/서버 오류(429/5xx)에 대해 지수 백오프로 재시도한다."""
	retry_codes = RETRY_STATUS_CODES if retry_status_codes is None else set(retry_status_codes)
	last_exc = None

	for attempt in range(max_retries):
		try:
			res = session.request(method=method, url=url, timeout=timeout, **kwargs)
			if res.status_code in retry_codes:
				raise requests.HTTPError(f'retryable status: {res.status_code}', response=res)
			return res
		except (requests.exceptions.Timeout, requests.exceptions.ConnectionError, requests.HTTPError) as exc:
			last_exc = exc
			if attempt >= max_retries - 1:
				raise
			sleep_s = (backoff_base * (2 ** attempt)) + random.uniform(0, 0.2)
			time.sleep(min(sleep_s, 5.0))

	if last_exc:
		raise last_exc
	raise RuntimeError('request failed without exception')


async def async_request_text_with_retry(session, method, url, *, semaphore=None, max_retries=3, backoff_base=0.6, retry_status_codes=None, **kwargs):
	"""aiohttp 요청을 재시도하며 응답 본문(text)을 반환한다."""
	retry_codes = RETRY_STATUS_CODES if retry_status_codes is None else set(retry_status_codes)
	last_exc = None

	for attempt in range(max_retries):
		try:
			if semaphore is not None:
				async with semaphore:
					async with session.request(method=method, url=url, **kwargs) as res:
						text = await res.text(encoding='utf-8', errors='ignore')
						if res.status in retry_codes:
							raise RuntimeError(f'retryable status: {res.status}')
						return text
			else:
				async with session.request(method=method, url=url, **kwargs) as res:
					text = await res.text(encoding='utf-8', errors='ignore')
					if res.status in retry_codes:
						raise RuntimeError(f'retryable status: {res.status}')
					return text
		except (aiohttp.ClientError, asyncio.TimeoutError, RuntimeError) as exc:
			last_exc = exc
			if attempt >= max_retries - 1:
				raise
			sleep_s = (backoff_base * (2 ** attempt)) + random.uniform(0, 0.2)
			await asyncio.sleep(min(sleep_s, 5.0))

	if last_exc:
		raise last_exc
	raise RuntimeError('async request failed without exception')


# ---------------------------
# 공통 상수 정의 (KOSHA)
# ---------------------------
SERVICE_KEY = st.secrets['SERVICE_KEY']


# ---------------------------
# 연간입고량/연간사용·판매량 코드 매핑
# ---------------------------
QUANTITY_CODE_MAPPING = {
	'01': '0.1미만',
	'02': '0.1~0.5',
	'03': '0.5~1.0',
	'04': '1~2.5',
	'05': '2.5~5.0',
	'06': '5~20',
	'07': '20~200',
	'08': '200~1,000',
	'09': '1,000~5,000',
	'10': '5,000이상',
}


# ---------------------------
# KOSHA 유해성 컬럼 순서 (makeResult 기준)
# ---------------------------
HAZARD_ORDER = [
	'#', '물질명칭', 'CAS No.', '결과없음',
	'발암성', '생식독성', '생식세포 변이원성', 'CMR',
	'급성 독성(경구)', '급성 독성(경피)', '급성 독성(흡입)', '급성 독성',
	'흡인 유해성', '피부 부식성/피부 자극성', '심한 눈 손상성/눈 자극성', '피부/눈 자극성',
	'호흡기 과민성', '피부 과민성', '피부/호흡기 과민성',
	'특정표적장기 독성(1회 노출)', '특정표적장기 독성(반복 노출)', '특정표적장기 독성',
	'급성 수생환경 유해성', '만성 수생환경 유해성', '수생환경 유해성',
	'폭발성 물질', '자기반응성 물질', '유기과산화물', '산화성 가스',
	'산화성 액체', '산화성 고체', '인화성 가스', '인화성 에어로졸',
	'인화성 액체', '인화성 고체', '인화성',
	'자연발화성 액체', '자연발화성 고체', '물반응성 물질', '고압가스',
	'자기발열성 물질', '금속부식성 물질', 'TWA', 'STEL', '증기압', '개정일',
	'관리대상유해물질', '특별관리물질', '작업환경측정대상물질', '특수건강진단대상물질',
	'노출기준설정물질', '허용기준설정물질', '금지물질', '제한물질',
	'인체급성유해성물질', '인체만성유해성물질', '생태유해성물질',
	'허가물질', '사고대비물질', '중점관리물질', '위험물', '독성가스',
	'[11번] 특정표적장기 독성(1회 노출)', '[11번] 특정표적장기 독성(반복 노출)',
]


# ---------------------------
# 등급 판별/집계용 상수
# ---------------------------
CMR_GRADE_PRIORITY = {'1A': 0, '1B': 1, '2': 2}
GRADE_PRIORITY = {'1': 0, '2': 1, '3': 2, '4': 3}

AGGREGATE_GROUPS = [
	('급성 독성', ['급성 독성(경구)', '급성 독성(경피)', '급성 독성(흡입)']),
	('피부/눈 자극성', ['피부 부식성/피부 자극성', '심한 눈 손상성/눈 자극성']),
	('피부/호흡기 과민성', ['호흡기 과민성', '피부 과민성']),
	('특정표적장기 독성', ['특정표적장기 독성(1회 노출)', '특정표적장기 독성(반복 노출)']),
	('수생환경 유해성', ['급성 수생환경 유해성', '만성 수생환경 유해성']),
	('인화성', ['인화성 가스', '인화성 에어로졸', '인화성 액체', '인화성 고체']),
]


# ---------------------------
# 엑셀 스타일 상수
# ---------------------------
default_font = Font(name='맑은 고딕', size=10, bold=False)
bold_font = Font(name='맑은 고딕', size=10, bold=True)

fill_header = PatternFill(fill_type='solid', fgColor='FF5B9BD5')
fill_label = PatternFill(fill_type='solid', fgColor='FFDDEBF7')
fill_data1 = PatternFill(fill_type='solid', fgColor='FFFFFF00')
fill_data2 = PatternFill(fill_type='solid', fgColor='FFFFC000')


# =========================================================
# 1) ICIS 조회/파싱 관련 함수
# =========================================================
def fetch_search_page(session):
	"""ICIS 검색 페이지를 먼저 호출해서 세션 쿠키를 안정적으로 확보한다."""
	res = request_with_retry(session, 'GET', SEARCH_PAGE_URL, headers=HEADERS, timeout=20)
	res.raise_for_status()
	return res.text


def fetch_search_results(session, company_name, page_no=1, search_year=None):
	"""업체명으로 ICIS JSON 검색 결과를 한 페이지 조회한다."""
	data = {
		'bplcNm': company_name,
		'pageNo': page_no,
	}
	if search_year:
		data['searchYear'] = search_year
	res = request_with_retry(session, 'POST', SEARCH_JSON_URL, headers=HEADERS, data=data, timeout=20)
	res.raise_for_status()
	return res.json()


def fetch_all_search_results(session, company_name, search_year=None):
	"""페이징을 끝까지 순회해서 업체 검색 결과를 전부 수집한다."""
	items = []
	page_no = 1
	while True:
		use_year = '2022' if search_year is None else search_year
		search_json = fetch_search_results(session, company_name, page_no=page_no, search_year=use_year)

		if search_json.get('result') != 'SUCCESS':
			return []

		page_items = search_json.get('list', [])
		total_count = int(search_json.get('totalCount', len(page_items)))
		if not page_items:
			break

		items.extend(page_items)
		if len(items) >= total_count:
			break

		page_no += 1

	return items


def fetch_detail_page(session, bplc_id, search_year=None):
	"""선택한 업체의 상세 페이지(화학물질 취급현황 포함 HTML)를 가져온다."""
	form_data = {
		'searchYear': '2022' if search_year is None else search_year,
		'bplcId': bplc_id,
		'streNo': '',
	}
	res = request_with_retry(session, 'POST', DETAIL_VIEW_URL, headers=HEADERS, data=form_data, timeout=30)
	res.raise_for_status()
	return res.text


def sanitize_filename(name):
	"""윈도우 파일명에서 사용할 수 없는 문자 제거."""
	cleaned = re.sub(r'[<>:"/\\|?*\n\r\t]+', '_', name)
	cleaned = cleaned.strip().rstrip('.')
	return cleaned or 'detail'


def extract_region_from_address(address):
	"""주소에서 지역명(시/군)만 추출해 중복 업체 파일명 구분에 사용한다."""
	if not address:
		return ''

	address = address.strip()
	match = re.search(r'([가-힣]+?(?:특별자치시|광역시|특별시|시|군))', address)
	if not match:
		return ''

	region = match.group(1)
	if region.endswith('특별자치시'):
		return region[:-6]
	if region.endswith('광역시') or region.endswith('특별시'):
		return region[:-3]
	return region[:-1]


def extract_company_name(html_text):
	"""상세 HTML에서 업체명을 안전하게 추출한다."""
	match = re.search(
		r'<th[^>]*>\s*업체명\s*</th>\s*<td[^>]*>(.*?)</td>',
		html_text,
		re.S | re.I,
	)
	if not match:
		return None

	company_name = re.sub(r'<[^>]+>', '', match.group(1))
	return unescape(company_name).strip()


def extract_section_html(html_text, section_title='3. 화학물질 취급현황'):
	"""특정 섹션(h4 제목 기준)만 잘라낸다."""
	section_search = re.search(
		rf'<h4[^>]*>\s*{re.escape(section_title)}\s*</h4>',
		html_text,
		re.S | re.I,
	)
	if not section_search:
		return None

	search_from = section_search.end()
	next_section = re.search(r'<h4[^>]*>\s*[0-9]+\.', html_text[search_from:], re.S | re.I)
	end_index = search_from + next_section.start() if next_section else len(html_text)
	return html_text[section_search.start():end_index]


def select_section_table(html_text, section_title='3. 화학물질 취급현황', table_index=2):
	"""섹션 내부에서 지정한 순번의 테이블(기본 2번째)을 선택한다."""
	section_html = extract_section_html(html_text, section_title=section_title)
	if not section_html:
		return None

	tables = re.findall(r'(<table[^>]*>.*?</table>)', section_html, re.S | re.I)
	if not tables:
		return None

	if 1 <= table_index <= len(tables):
		return tables[table_index - 1]
	return tables[-1]


class TableHTMLParser(HTMLParser):
	"""테이블 셀 텍스트를 2차원 리스트로 추출하는 단순 파서."""

	def __init__(self):
		super().__init__()
		self.in_table = False
		self.table_depth = 0
		self.in_tr = False
		self.in_cell = False
		self.current_cell = ''
		self.current_row = []
		self.rows = []

	def handle_starttag(self, tag, attrs):
		if tag == 'table':
			if self.in_table:
				self.table_depth += 1
			else:
				self.in_table = True
				self.table_depth = 1

		if self.in_table and tag == 'tr':
			self.in_tr = True
			self.current_row = []

		if self.in_tr and tag in ('td', 'th'):
			self.in_cell = True
			self.current_cell = ''

	def handle_data(self, data):
		if self.in_cell:
			self.current_cell += data

	def handle_endtag(self, tag):
		if tag in ('td', 'th') and self.in_cell:
			text = unescape(self.current_cell).strip()
			self.current_row.append(' '.join(text.split()))
			self.in_cell = False
		elif tag == 'tr' and self.in_tr:
			if self.current_row:
				self.rows.append(self.current_row)
			self.in_tr = False
		elif tag == 'table' and self.in_table:
			self.table_depth -= 1
			if self.table_depth <= 0:
				self.in_table = False


def parse_table_html(table_html):
	"""HTML 문자열에서 테이블을 파싱해 행 리스트로 반환한다.

	1) pandas.read_html 우선 시도(HTML 구조 변경에 상대적으로 강함)
	2) 실패 시 기존 커스텀 파서로 fallback
	"""
	try:
		tables = pd.read_html(io.StringIO(table_html), header=None)
		if tables:
			df = tables[0].fillna('')
			rows = []
			for _, rec in df.iterrows():
				row = [' '.join(str(v).split()) for v in rec.tolist()]
				if any(cell for cell in row):
					rows.append(row)
			return rows
	except Exception:
		pass

	parser = TableHTMLParser()
	parser.feed(table_html)
	return parser.rows[2:] if len(parser.rows) > 2 else parser.rows


def _is_header_only_row(row):
	"""헤더 키워드만 반복되는 잡음 행인지 판단한다."""
	header_keywords = [
		'물질명칭', 'CAS', 'CAS No', 'CAS No.', '제품', '제품명', '인체등유해성물질',
		'제한물질2', '금지물질2', '허가물질2', '사고대비물질2', '중점관리물질2',
		'금지·허가물질2', '노출·허용기준물질2', '직업환경측정물질등2', '위험물2',
		'독성가스2', '연간입고량', '연간사용·판매량',
	]
	normalized_keywords = [kw.lower().replace('.', '').replace('·', ' ') for kw in header_keywords]
	cells = [cell.strip().lower().replace('.', '').replace('·', ' ') for cell in row if cell.strip()]
	if not cells:
		return False
	return all(any(nk == cell or nk in cell or cell in nk for nk in normalized_keywords) for cell in cells)


def _convert_quantity_code(value):
	"""01~10 코드를 사람 읽기 범위 문자열로 바꾼다."""
	if not isinstance(value, str):
		return value
	code = value.strip()
	return QUANTITY_CODE_MAPPING.get(code, value)


def clean_rows(rows):
	"""헤더 위치를 기준으로 데이터 구간만 정리한다."""
	if not rows:
		return rows

	header_idx = None
	for i, row in enumerate(rows):
		joined = ' '.join(row)
		if 'CAS' in joined or 'CAS No' in joined or 'CAS No.' in joined:
			header_idx = i
			break

	if header_idx is not None:
		rows = rows[header_idx:]
		if len(rows) >= 2 and _is_header_only_row(rows[1]):
			return [rows[0]] + rows[2:]
		return rows

	return rows


def workbook_to_bytes(rows):
	"""ICIS 결과를 기존 형식(77열) 엑셀 바이트로 만든다."""
	wb = Workbook()
	ws = wb.active

	# 기존 규칙을 그대로 유지하기 위해 헤더를 고정 목록으로 둔다.
	full_headers = [
		'#', '물질명칭', 'CAS No.', '결과없음', '발암성', '생식독성', '생식세포 변이원성', 'CMR',
		'급성 독성(경구)', '급성 독성(경피)', '급성 독성(흡입)', '급성 독성', '흡인 유해성',
		'피부 부식성/피부 자극성', '심한 눈 손상성/눈 자극성', '피부/눈 자극성', '호흡기 과민성',
		'피부 과민성', '피부/호흡기 과민성', '특정표적장기 독성(1회 노출)', '특정표적장기 독성(반복 노출)',
		'특정표적장기 독성', '급성 수생환경 유해성', '만성 수생환경 유해성', '수생환경 유해성',
		'폭발성 물질', '자기반응성 물질', '유기과산화물', '산화성 가스', '산화성 액체',
		'산화성 고체', '인화성 가스', '인화성 에어로졸', '인화성 액체', '인화성 고체',
		'인화성', '자연발화성 액체', '자연발화성 고체', '물반응성 물질', '고압가스',
		'자기발열성 물질', '금속부식성 물질', 'TWA', 'STEL', '증기압', '개정일',
		'관리대상유해물질', '특별관리물질', '작업환경측정대상물질', '특수건강진단대상물질',
		'노출기준설정물질', '허용기준설정물질', '금지물질', '제한물질', '인체급성유해성물질',
		'인체만성유해성물질', '생태유해성물질', '허가물질', '사고대비물질', '중점관리물질',
		'위험물', '독성가스', '인체등유해성물질', '제한물질2', '금지물질2', '허가물질2',
		'사고대비물질2', '중점관리물질2', '금지·허가물질2', '노출·허용기준물질2',
		'직업환경측정물질등2', '위험물2', '독성가스2', '연간입고량', '연간사용·판매량',
		'[11번] 특정표적장기 독성(1회 노출)', '[11번] 특정표적장기 독성(반복 노출)',
	]

	header_font = Font(name='맑은 고딕', size=10, bold=False)
	header_font_bold = Font(name='맑은 고딕', size=10, bold=True)
	header_fill = PatternFill(start_color='FF5B9BD5', end_color='FF5B9BD5', fill_type='solid')
	header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

	bold_header_cols = {5, 6, 8, 12, 16, 22, 25, 36, 75}
	header_comments = {
		10: '경피 급성독성 화학물질을 쓸 때 차폐 및 보호장구 매우 중요함',
		11: '흡입시 사망 가능가스를 쓸 때 차폐/국소배기장치/환기 및 보호장구 매우 중요함',
		13: '흡입유해성이 큰 화학물질을 쓸 때 차폐/국소배기장치/환기 및 보호장구 매우 중요함',
		24: '만성 수생환경 유해성이 있는 물질이 우수관 등을 통해서 하천으로 흘러들면 하천생태계에 큰 영향을 줌',
		32: '인화성 가스를 쓰는 공정에서 접지 불량 등 발견되면 심각한 화재위험이 됨',
	}

	data_font = Font(name='맑은 고딕', size=10)
	fill_abc = PatternFill(start_color='FFDDEBF7', end_color='FFDDEBF7', fill_type='solid')
	fill_d_bj = PatternFill(start_color='FFFFFF00', end_color='FFFFFF00', fill_type='solid')
	fill_bk_bw = PatternFill(start_color='FFFFC000', end_color='FFFFC000', fill_type='solid')
	fill_bx_by = PatternFill(start_color='FFFFFF00', end_color='FFFFFF00', fill_type='solid')

	alignment_default_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
	alignment_default_left = Alignment(horizontal='left', vertical='center', wrap_text=True)

	thin_border = Border(
		left=Side(style='thin', color='000000'),
		right=Side(style='thin', color='000000'),
		top=Side(style='thin', color='000000'),
		bottom=Side(style='thin', color='000000'),
	)

	# 헤더 작성
	for c_idx, header_text in enumerate(full_headers, start=1):
		cell = ws.cell(row=1, column=c_idx, value=header_text)
		cell.font = header_font_bold if c_idx in bold_header_cols else header_font
		cell.fill = header_fill
		cell.alignment = header_alignment
		cell.border = thin_border
		if c_idx in header_comments:
			cell.comment = Comment(header_comments[c_idx], 'author')

	ws.row_dimensions[1].height = 52.2

	# 데이터 행 작성
	if rows and len(rows) > 1:
		for r_idx, row in enumerate(rows[1:], start=2):
			for c_idx in range(1, 78):
				cell = ws.cell(row=r_idx, column=c_idx)
				cell.font = data_font
				cell.border = thin_border

				# 열 그룹별 배경색
				if 1 <= c_idx <= 3:
					cell.fill = fill_abc
				elif 4 <= c_idx <= 62:
					cell.fill = fill_d_bj
				elif 63 <= c_idx <= 75:
					cell.fill = fill_bk_bw
				else:
					cell.fill = fill_bx_by

				# 열별 값 배치
				if c_idx == 1:
					cell.value = r_idx - 1
					cell.alignment = alignment_default_center
				elif c_idx == 2 and len(row) > 0:
					cell.value = row[0]
					cell.alignment = alignment_default_left
				elif c_idx == 3 and len(row) > 1:
					cell.value = str(row[1]) if row[1] is not None else None
					cell.alignment = alignment_default_center
					cell.number_format = '@'
				elif 63 <= c_idx <= 77 and len(row) > (c_idx - 61):
					cell.value = row[c_idx - 61]
					cell.alignment = alignment_default_center
				else:
					cell.value = None
					cell.alignment = alignment_default_center

	# 열 너비 고정
	column_widths = {
		'A': 5.0, 'B': 31.0, 'C': 13.0, 'D': 12.0, 'E': 10.0, 'F': 10.0, 'G': 7.09765625,
		'H': 9.0, 'I': 8.296875, 'J': 8.8984375, 'K': 8.3984375, 'L': 8.296875, 'M': 8.59765625,
		'N': 10.19921875, 'O': 10.8984375, 'P': 11.0, 'Q': 8.59765625, 'R': 10.0, 'S': 10.0,
		'T': 12.0, 'U': 12.0, 'V': 12.0, 'W': 9.5, 'X': 9.5, 'Y': 9.0, 'Z': 7.0,
		'AA': 9.0, 'AB': 9.0, 'AC': 7.0, 'AD': 7.0, 'AE': 7.0, 'AF': 7.0, 'AG': 8.0,
		'AH': 7.0, 'AI': 7.0, 'AJ': 8.0, 'AK': 9.0, 'AL': 9.0, 'AM': 9.0, 'AN': 9.0,
		'AO': 9.0, 'AP': 9.0, 'AQ': 9.0, 'AR': 9.0, 'AS': 9.5, 'AT': 9.0, 'AU': 9.0,
		'AV': 9.0, 'AW': 9.0, 'AX': 9.0, 'AY': 9.0, 'AZ': 9.0, 'BA': 8.0, 'BB': 8.0,
		'BC': 9.0, 'BD': 9.0, 'BE': 8.0, 'BF': 8.0, 'BG': 8.0, 'BH': 8.0, 'BI': 8.0,
		'BJ': 8.0, 'BK': 9.0, 'BL': 9.0, 'BM': 9.0, 'BN': 9.0, 'BO': 8.0, 'BP': 8.0,
		'BQ': 9.0, 'BR': 9.0, 'BS': 9.0, 'BT': 9.0, 'BU': 9.0, 'BV': 9.5, 'BW': 9.5,
		'BX': 15.0, 'BY': 15.0,
	}
	for col_letter, width in column_widths.items():
		ws.column_dimensions[col_letter].width = width

	ws.sheet_view.zoomScale = 80

	buffer = io.BytesIO()
	wb.save(buffer)
	buffer.seek(0)
	return buffer.getvalue()


def create_excel_bytes_for_company(session, company_name, item, search_year=None):
	"""ICIS 데이터로 업체별 기본 엑셀(아직 KOSHA 미반영)을 만든다."""
	detail_html = fetch_detail_page(session, item['bplcId'], search_year=search_year)
	company_name_full = extract_company_name(detail_html) or item.get('bplcNm', company_name)

	target_table_html = select_section_table(
		detail_html,
		section_title='3. 화학물질 취급현황',
		table_index=2,
	)
	if not target_table_html:
		raise RuntimeError('화학물질 취급현황 섹션의 두번째 표를 찾을 수 없습니다.')

	rows = parse_table_html(target_table_html)
	rows = clean_rows(rows)
	if not rows:
		raise RuntimeError('표에서 데이터를 찾을 수 없습니다.')

	first_join = ' '.join(rows[0]) if rows else ''
	data_rows = rows[1:] if ('CAS' in first_join or 'CAS No' in first_join or 'CAS No.' in first_join) else rows

	# BK~BY로 들어갈 ICIS 원본 데이터 헤더
	header = [
		'물질명칭', 'CAS No.',
		'인체등유해성물질', '제한물질2', '금지물질2', '허가물질2', '사고대비물질2', '중점관리물질2',
		'금지·허가물질2', '노출·허용기준물질2', '직업환경측정물질등2', '위험물2', '독성가스2',
		'연간입고량', '연간사용·판매량', '[11번] 특정표적장기 독성(1회 노출)', '[11번] 특정표적장기 독성(반복 노출)',
	]

	# 연간입고량/연간사용·판매량 코드 변환
	quantity_columns = {13, 14}
	converted_rows = []
	for row in data_rows:
		converted_row = []
		for idx, cell in enumerate(row):
			if idx in quantity_columns:
				converted_row.append(_convert_quantity_code(cell))
			else:
				converted_row.append(cell)
		converted_rows.append(converted_row)

	final_rows = [header] + converted_rows
	return company_name_full, workbook_to_bytes(final_rows)


def search_companies(keywords, search_year=None):
	"""검색어 리스트로 업체 목록(중복명 구분용 지역 포함)을 만든다."""
	session = requests.Session()
	fetch_search_page(session)

	search_results = []
	fetched_entries = []
	missing_keywords = []
	for keyword in keywords:
		items = fetch_all_search_results(session, keyword, search_year=search_year)
		if items:
			for item in items:
				fetched_entries.append((keyword, item))
		else:
			missing_keywords.append(keyword)

	company_names = [item.get('bplcNm', '') for _, item in fetched_entries if item.get('bplcNm')]
	duplicate_counts = Counter(company_names)

	for keyword, item in fetched_entries:
		company_name = item.get('bplcNm', '')
		company_name_key = company_name
		company_address = item.get('locplcAdres', '') or item.get('bplcAdres', '')
		region = extract_region_from_address(company_address) or ''

		search_results.append({
			'index': len(search_results),
			'bplcId': item.get('bplcId', ''),
			'company_name': company_name,
			'region': region,
			'use_region_in_filename': duplicate_counts[company_name_key] > 1,
			'search_keyword': keyword,
		})

	for keyword in missing_keywords:
		search_results.append({
			'index': len(search_results),
			'company_name': f'검색 결과 없음: {keyword}',
			'region': '',
			'use_region_in_filename': False,
			'search_keyword': keyword,
		})

	return search_results


def zip_files(files_dict, selected_filenames):
	"""선택된 결과 파일들을 ZIP으로 묶는다."""
	zip_bytes = io.BytesIO()
	with zipfile.ZipFile(zip_bytes, 'w', zipfile.ZIP_DEFLATED) as archive:
		for filename in selected_filenames:
			data = files_dict.get(filename)
			if data is not None:
				archive.writestr(filename, data)
	zip_bytes.seek(0)
	return zip_bytes.getvalue()


# =========================================================
# 2) KOSHA 조회/채움 관련 함수
# =========================================================
def extract_cmr_grades(val):
	if not isinstance(val, str) or not val.strip():
		return []
	grades = []
	for part in re.split(r'[\n|,]+', val):
		m = re.match(r'^\s*(1A|1B|2)\b', part.strip())
		if m:
			grades.append(m.group(1))
	return grades


def extract_grades(val):
	if not isinstance(val, str) or not val.strip():
		return []
	grades = []
	for part in re.split(r'[\n|,]+', val):
		m = re.match(r'^\s*(1|2|3|4)\b', part.strip())
		if m:
			grades.append(m.group(1))
	return grades


def get_highest_cmr_grade(grades):
	valid = [g for g in grades if g in CMR_GRADE_PRIORITY]
	if not valid:
		return None
	return min(valid, key=lambda g: CMR_GRADE_PRIORITY[g])


def get_highest_grade(grades):
	valid = [g for g in grades if g in GRADE_PRIORITY]
	if not valid:
		return None
	return min(valid, key=lambda g: GRADE_PRIORITY[g])


def compute_aggregate_grade(result, source_keys):
	grades = []
	for k in source_keys:
		grades.extend(extract_grades(result.get(k, '')))
	return get_highest_grade(grades)


def query_twa_stel(session, service_key, chem_id):
	"""상세08에서 TWA/STEL을 파싱한다."""
	try:
		res = request_with_retry(
			session,
			'GET',
			'https://msds.kosha.or.kr/openapi/service/msdschem/chemdetail08',
			params={'serviceKey': service_key, 'chemId': chem_id},
			timeout=10,
		)
		res.encoding = 'utf-8'
		root = ET.fromstring(res.text)

		twa = ''
		stel = ''
		for item in root.findall('.//item'):
			name_twa = item.findtext('msdsItemNameKor')
			if name_twa == '국내규정':
				detail = item.findtext('itemDetail')
				if detail:
					parts = [p.strip() for p in detail.split('|') if p.strip()]
					for part in parts:
						if part.startswith('TWA'):
							match = re.search(r'TWA\s*[:]?\s*([\d\.]+\s*ppm(?:\([^)]*\))?)', part)
							if match:
								twa = match.group(1).strip()
						elif part.startswith('STEL'):
							match = re.search(r'STEL\s*[:]?\s*([\d\.]+\s*ppm(?:\([^)]*\))?)', part)
							if match:
								stel = match.group(1).strip()

		return twa, stel
	except Exception:
		return '', ''


def query_vapor_pressure(session, service_key, chem_id):
	"""상세09에서 증기압을 파싱한다."""
	try:
		res = request_with_retry(
			session,
			'GET',
			'https://msds.kosha.or.kr/openapi/service/msdschem/chemdetail09',
			params={'serviceKey': service_key, 'chemId': chem_id},
			timeout=10,
		)
		res.encoding = 'utf-8'
		root = ET.fromstring(res.text)

		for item in root.findall('.//item'):
			name = item.findtext('msdsItemNameKor')
			if name and name.strip() == '증기압':
				detail = html.unescape(item.findtext('itemDetail') or '')
				if detail.strip():
					cleaned = ''.join(detail.split())
					cleaned = re.split(r'\|+|※+', cleaned)[0]
					return cleaned
	except Exception:
		pass
	return ''


def query_revision_date(session, service_key, chem_id):
	"""상세16에서 최종 개정일자를 가져온다."""
	try:
		res = request_with_retry(
			session,
			'GET',
			'https://msds.kosha.or.kr/openapi/service/msdschem/chemdetail16',
			params={'serviceKey': service_key, 'chemId': chem_id},
			timeout=10,
		)
		res.encoding = 'utf-8'
		root = ET.fromstring(res.text)
		for item in root.findall('.//item'):
			if item.findtext('msdsItemNameKor') == '최종 개정일자':
				detail = item.findtext('itemDetail')
				return detail.strip() if detail else ''
	except Exception:
		return ''
	return ''


def has_keyword(detail, keyword):
	if not detail:
		return False
	items = [d.strip() for d in detail.split('|')]
	return any(keyword in item for item in items)


def query_detail15(session, service_key, chem_id):
	"""상세15 규제/관리 플래그(▣ 대상)를 판정한다."""
	try:
		res = request_with_retry(
			session,
			'GET',
			'https://msds.kosha.or.kr/openapi/service/msdschem/chemdetail15',
			params={'serviceKey': service_key, 'chemId': chem_id},
			timeout=10,
		)
		res.encoding = 'utf-8'
		root = ET.fromstring(res.text)

		result = {
			'관리대상유해물질': False, '특별관리물질': False, '작업환경측정대상물질': False, '특수건강진단대상물질': False,
			'노출기준설정물질': False, '허용기준설정물질': False, '금지물질': False, '제한물질': False,
			'허가물질': False, '사고대비물질': False, '중점관리물질': False, '위험물': False, '독성가스': False,
			'인체급성유해성물질': False, '인체만성유해성물질': False, '생태유해성물질': False,
		}

		for item in root.findall('.//item'):
			code = item.findtext('msdsItemCode')
			detail = item.findtext('itemDetail') or ''

			if code in ['002', 'O04', 'O12', 'O06']:
				if has_keyword(detail, '금지물질'):
					result['금지물질'] = True
				if has_keyword(detail, '제한물질'):
					result['제한물질'] = True
				if has_keyword(detail, '허가물질'):
					result['허가물질'] = True
				if has_keyword(detail, '사고대비물질'):
					result['사고대비물질'] = True
				if has_keyword(detail, '중점관리물질'):
					result['중점관리물질'] = True
				if has_keyword(detail, '위험물'):
					result['위험물'] = True
				if has_keyword(detail, '독성가스'):
					result['독성가스'] = True

			if code == 'O02':
				if has_keyword(detail, '관리대상유해물질'):
					result['관리대상유해물질'] = True
				if has_keyword(detail, '특별관리물질'):
					result['특별관리물질'] = True
				if has_keyword(detail, '특수건강진단대상물질'):
					result['특수건강진단대상물질'] = True
				if has_keyword(detail, '작업환경측정대상물질'):
					result['작업환경측정대상물질'] = True
				if has_keyword(detail, '노출기준설정물질'):
					result['노출기준설정물질'] = True
				if has_keyword(detail, '허용기준설정물질'):
					result['허용기준설정물질'] = True

			elif code in ['O04', 'O12']:
				if has_keyword(detail, '인체급성유해성물질'):
					result['인체급성유해성물질'] = True
				if has_keyword(detail, '인체만성유해성물질'):
					result['인체만성유해성물질'] = True
				if has_keyword(detail, '생태유해성물질'):
					result['생태유해성물질'] = True

		return result
	except Exception:
		return {
			'관리대상유해물질': False, '특별관리물질': False, '작업환경측정대상물질': False, '특수건강진단대상물질': False,
			'노출기준설정물질': False, '허용기준설정물질': False, '금지물질': False, '제한물질': False,
			'허가물질': False, '사고대비물질': False, '중점관리물질': False, '위험물': False, '독성가스': False,
			'인체급성유해성물질': False, '인체만성유해성물질': False, '생태유해성물질': False,
		}


def query_detail11(session, service_key, chem_id):
	"""상세11에서 [11번] 특정표적장기 독성 문구를 가져온다."""
	try:
		res = request_with_retry(
			session,
			'GET',
			'https://msds.kosha.or.kr/openapi/service/msdschem/chemdetail11',
			params={'serviceKey': service_key, 'chemId': chem_id},
			timeout=10,
		)
		res.encoding = 'utf-8'
		root = ET.fromstring(res.text)

		stot_single = ''
		stot_repeated = ''
		for item in root.findall('.//item'):
			code = item.findtext('msdsItemCode')
			detail = html.unescape(item.findtext('itemDetail') or '')
			if detail.strip() == '자료없음':
				continue
			if code == 'K0418':
				stot_single = detail.strip()
			elif code == 'K0420':
				stot_repeated = detail.strip()

		return stot_single, stot_repeated
	except Exception:
		return '', ''


def process_single_chemical(idx, row, session, service_key):
	"""한 물질(한 행)에 대해 KOSHA API 조회 후 결과 dict를 만든다."""
	cas_raw = str(row[2]).strip()
	cas = '' if cas_raw.lower() in ['nan', 'none', 'null'] else cas_raw

	id_num = row[0]
	name = row[1]
	result = {'#': id_num, '물질명칭': name, 'CAS No.': cas}
	unknown_cols = set()

	# CAS No. 예외 규칙 유지
	if not cas:
		result['결과없음'] = '영업비밀'
		return idx, result, unknown_cols
	if '심의중' in cas:
		result['결과없음'] = '심의중'
		return idx, result, unknown_cols

	try:
		# 1) CAS로 chemId 검색
		res_id = request_with_retry(
			session,
			'GET',
			'https://msds.kosha.or.kr/openapi/service/msdschem/chemlist',
			params={'serviceKey': service_key, 'searchWrd': cas, 'searchCnd': 1},
			timeout=10,
		)
		res_id.encoding = 'utf-8'
		chem_id = ET.fromstring(res_id.text).findtext('.//chemId')

		if not chem_id:
			result['결과없음'] = '공단 MSDS 없음'
			return idx, result, unknown_cols

		# 2) 상세02에서 유해성 본문 파싱
		res_detail = request_with_retry(
			session,
			'GET',
			'https://msds.kosha.or.kr/openapi/service/msdschem/chemdetail02',
			params={'serviceKey': service_key, 'chemId': chem_id},
			timeout=10,
		)
		res_detail.encoding = 'utf-8'
		root = ET.fromstring(res_detail.text)

		b02_detail = next(
			(
				item.findtext('itemDetail')
				for item in root.findall('.//item')
				if item.findtext('msdsItemCode') == 'B02'
			),
			None,
		)

		if b02_detail is None or b02_detail.strip() == '' or b02_detail.strip() == '자료없음':
			result['결과없음'] = '자료 없음'
		else:
			merged = defaultdict(list)
			inhalation_labels = [
				'급성 독성(흡입)', '급성 독성(흡입: 가스)',
				'급성 독성(흡입: 분진/미스트)', '급성 독성(흡입: 증기)',
			]
			inhalation_entries = []
			cmr_map = {'발암성': [], '생식독성': [], '생식세포 변이원성': []}

			for entry in b02_detail.split('|'):
				if ':' in entry and '자료없음' not in entry:
					k, v = map(str.strip, entry.rsplit(':', 1))
					v = v.replace('구분', '').strip()

					if k in inhalation_labels:
						label = k.replace('급성 독성(', '').replace(')', '')
						inhalation_entries.append(f'{v}({label})')
					elif k in cmr_map:
						if not (cas == '64-17-5' and k == '발암성'):
							cmr_map[k].append(v)
							if v not in merged[k]:
								merged[k].append(v)
					else:
						if v not in merged[k]:
							merged[k].append(v)

			if inhalation_entries:
				result['급성 독성(흡입)'] = '\n'.join(inhalation_entries)

			for k, v_list in merged.items():
				if cas == '64-17-5' and k == '발암성':
					continue
				result[k] = '\n'.join(v_list)
				if k not in HAZARD_ORDER:
					unknown_cols.add(k)

			cmr_grades = []
			for values in cmr_map.values():
				for v in values:
					cmr_grades.extend(extract_cmr_grades(v))
			highest_grade = get_highest_cmr_grade(cmr_grades)
			if highest_grade:
				result['CMR'] = highest_grade

			for agg_col, source_keys in AGGREGATE_GROUPS:
				grade = compute_aggregate_grade(result, source_keys)
				if grade:
					result[agg_col] = grade

		# 3) 기타 상세 정보 병합
		result['TWA'], result['STEL'] = query_twa_stel(session, service_key, chem_id)
		result['증기압'] = query_vapor_pressure(session, service_key, chem_id)
		result['개정일'] = query_revision_date(session, service_key, chem_id)

		res_detail15 = query_detail15(session, service_key, chem_id)
		if res_detail15['관리대상유해물질']:
			result['관리대상유해물질'] = '▣'
		if res_detail15['특별관리물질']:
			result['특별관리물질'] = '▣'
		if res_detail15['특수건강진단대상물질']:
			result['특수건강진단대상물질'] = '▣'
		if res_detail15['인체급성유해성물질']:
			result['인체급성유해성물질'] = '▣'
		if res_detail15['인체만성유해성물질']:
			result['인체만성유해성물질'] = '▣'
		if res_detail15['생태유해성물질']:
			result['생태유해성물질'] = '▣'
		if res_detail15['작업환경측정대상물질']:
			result['작업환경측정대상물질'] = '▣'
		if res_detail15['노출기준설정물질']:
			result['노출기준설정물질'] = '▣'
		if res_detail15['허용기준설정물질']:
			result['허용기준설정물질'] = '▣'
		if res_detail15['금지물질']:
			result['금지물질'] = '▣'
		if res_detail15['제한물질']:
			result['제한물질'] = '▣'
		if res_detail15['허가물질']:
			result['허가물질'] = '▣'
		if res_detail15['사고대비물질']:
			result['사고대비물질'] = '▣'
		if res_detail15['중점관리물질']:
			result['중점관리물질'] = '▣'
		if res_detail15['위험물']:
			result['위험물'] = '▣'
		if res_detail15['독성가스']:
			result['독성가스'] = '▣'

		stot_single, stot_repeated = query_detail11(session, service_key, chem_id)
		if stot_single:
			result['[11번] 특정표적장기 독성(1회 노출)'] = stot_single
		if stot_repeated:
			result['[11번] 특정표적장기 독성(반복 노출)'] = stot_repeated

	except Exception as e:
		result['결과없음'] = f'조회 오류: {str(e)}'

	return idx, result, unknown_cols


async def _query_twa_stel_async(session, semaphore, service_key, chem_id):
	try:
		text = await async_request_text_with_retry(
			session,
			'GET',
			'https://msds.kosha.or.kr/openapi/service/msdschem/chemdetail08',
			semaphore=semaphore,
			params={'serviceKey': service_key, 'chemId': chem_id},
		)
		root = ET.fromstring(text)
		twa = ''
		stel = ''
		for item in root.findall('.//item'):
			name_twa = item.findtext('msdsItemNameKor')
			if name_twa == '국내규정':
				detail = item.findtext('itemDetail')
				if detail:
					parts = [p.strip() for p in detail.split('|') if p.strip()]
					for part in parts:
						if part.startswith('TWA'):
							match = re.search(r'TWA\s*[:]?\s*([\d\.]+\s*ppm(?:\([^)]*\))?)', part)
							if match:
								twa = match.group(1).strip()
						elif part.startswith('STEL'):
							match = re.search(r'STEL\s*[:]?\s*([\d\.]+\s*ppm(?:\([^)]*\))?)', part)
							if match:
								stel = match.group(1).strip()
		return twa, stel
	except Exception:
		return '', ''


async def _query_vapor_pressure_async(session, semaphore, service_key, chem_id):
	try:
		text = await async_request_text_with_retry(
			session,
			'GET',
			'https://msds.kosha.or.kr/openapi/service/msdschem/chemdetail09',
			semaphore=semaphore,
			params={'serviceKey': service_key, 'chemId': chem_id},
		)
		root = ET.fromstring(text)
		for item in root.findall('.//item'):
			name = item.findtext('msdsItemNameKor')
			if name and name.strip() == '증기압':
				detail = html.unescape(item.findtext('itemDetail') or '')
				if detail.strip():
					cleaned = ''.join(detail.split())
					cleaned = re.split(r'\|+|※+', cleaned)[0]
					return cleaned
	except Exception:
		pass
	return ''


async def _query_revision_date_async(session, semaphore, service_key, chem_id):
	try:
		text = await async_request_text_with_retry(
			session,
			'GET',
			'https://msds.kosha.or.kr/openapi/service/msdschem/chemdetail16',
			semaphore=semaphore,
			params={'serviceKey': service_key, 'chemId': chem_id},
		)
		root = ET.fromstring(text)
		for item in root.findall('.//item'):
			if item.findtext('msdsItemNameKor') == '최종 개정일자':
				detail = item.findtext('itemDetail')
				return detail.strip() if detail else ''
	except Exception:
		return ''
	return ''


async def _query_detail15_async(session, semaphore, service_key, chem_id):
	default = {
		'관리대상유해물질': False, '특별관리물질': False, '작업환경측정대상물질': False, '특수건강진단대상물질': False,
		'노출기준설정물질': False, '허용기준설정물질': False, '금지물질': False, '제한물질': False,
		'허가물질': False, '사고대비물질': False, '중점관리물질': False, '위험물': False, '독성가스': False,
		'인체급성유해성물질': False, '인체만성유해성물질': False, '생태유해성물질': False,
	}
	try:
		text = await async_request_text_with_retry(
			session,
			'GET',
			'https://msds.kosha.or.kr/openapi/service/msdschem/chemdetail15',
			semaphore=semaphore,
			params={'serviceKey': service_key, 'chemId': chem_id},
		)
		root = ET.fromstring(text)
		result = dict(default)

		for item in root.findall('.//item'):
			code = item.findtext('msdsItemCode')
			detail = item.findtext('itemDetail') or ''

			if code in ['002', 'O04', 'O12', 'O06']:
				if has_keyword(detail, '금지물질'):
					result['금지물질'] = True
				if has_keyword(detail, '제한물질'):
					result['제한물질'] = True
				if has_keyword(detail, '허가물질'):
					result['허가물질'] = True
				if has_keyword(detail, '사고대비물질'):
					result['사고대비물질'] = True
				if has_keyword(detail, '중점관리물질'):
					result['중점관리물질'] = True
				if has_keyword(detail, '위험물'):
					result['위험물'] = True
				if has_keyword(detail, '독성가스'):
					result['독성가스'] = True

			if code == 'O02':
				if has_keyword(detail, '관리대상유해물질'):
					result['관리대상유해물질'] = True
				if has_keyword(detail, '특별관리물질'):
					result['특별관리물질'] = True
				if has_keyword(detail, '특수건강진단대상물질'):
					result['특수건강진단대상물질'] = True
				if has_keyword(detail, '작업환경측정대상물질'):
					result['작업환경측정대상물질'] = True
				if has_keyword(detail, '노출기준설정물질'):
					result['노출기준설정물질'] = True
				if has_keyword(detail, '허용기준설정물질'):
					result['허용기준설정물질'] = True

			elif code in ['O04', 'O12']:
				if has_keyword(detail, '인체급성유해성물질'):
					result['인체급성유해성물질'] = True
				if has_keyword(detail, '인체만성유해성물질'):
					result['인체만성유해성물질'] = True
				if has_keyword(detail, '생태유해성물질'):
					result['생태유해성물질'] = True

		return result
	except Exception:
		return default


async def _query_detail11_async(session, semaphore, service_key, chem_id):
	try:
		text = await async_request_text_with_retry(
			session,
			'GET',
			'https://msds.kosha.or.kr/openapi/service/msdschem/chemdetail11',
			semaphore=semaphore,
			params={'serviceKey': service_key, 'chemId': chem_id},
		)
		root = ET.fromstring(text)
		stot_single = ''
		stot_repeated = ''
		for item in root.findall('.//item'):
			code = item.findtext('msdsItemCode')
			detail = html.unescape(item.findtext('itemDetail') or '')
			if detail.strip() == '자료없음':
				continue
			if code == 'K0418':
				stot_single = detail.strip()
			elif code == 'K0420':
				stot_repeated = detail.strip()
		return stot_single, stot_repeated
	except Exception:
		return '', ''


async def process_single_chemical_async(idx, row, session, service_key, semaphore):
	"""한 물질을 비동기로 조회해 결과 dict를 만든다."""
	cas_raw = str(row[2]).strip()
	cas = '' if cas_raw.lower() in ['nan', 'none', 'null'] else cas_raw

	id_num = row[0]
	name = row[1]
	result = {'#': id_num, '물질명칭': name, 'CAS No.': cas}
	unknown_cols = set()

	if not cas:
		result['결과없음'] = '영업비밀'
		return idx, result, unknown_cols
	if '심의중' in cas:
		result['결과없음'] = '심의중'
		return idx, result, unknown_cols

	try:
		res_id_text = await async_request_text_with_retry(
			session,
			'GET',
			'https://msds.kosha.or.kr/openapi/service/msdschem/chemlist',
			semaphore=semaphore,
			params={'serviceKey': service_key, 'searchWrd': cas, 'searchCnd': 1},
		)
		chem_id = ET.fromstring(res_id_text).findtext('.//chemId')

		if not chem_id:
			result['결과없음'] = '공단 MSDS 없음'
			return idx, result, unknown_cols

		res_detail_text = await async_request_text_with_retry(
			session,
			'GET',
			'https://msds.kosha.or.kr/openapi/service/msdschem/chemdetail02',
			semaphore=semaphore,
			params={'serviceKey': service_key, 'chemId': chem_id},
		)
		root = ET.fromstring(res_detail_text)

		b02_detail = next(
			(
				item.findtext('itemDetail')
				for item in root.findall('.//item')
				if item.findtext('msdsItemCode') == 'B02'
			),
			None,
		)

		if b02_detail is None or b02_detail.strip() == '' or b02_detail.strip() == '자료없음':
			result['결과없음'] = '자료 없음'
		else:
			merged = defaultdict(list)
			inhalation_labels = [
				'급성 독성(흡입)', '급성 독성(흡입: 가스)',
				'급성 독성(흡입: 분진/미스트)', '급성 독성(흡입: 증기)',
			]
			inhalation_entries = []
			cmr_map = {'발암성': [], '생식독성': [], '생식세포 변이원성': []}

			for entry in b02_detail.split('|'):
				if ':' in entry and '자료없음' not in entry:
					k, v = map(str.strip, entry.rsplit(':', 1))
					v = v.replace('구분', '').strip()

					if k in inhalation_labels:
						label = k.replace('급성 독성(', '').replace(')', '')
						inhalation_entries.append(f'{v}({label})')
					elif k in cmr_map:
						if not (cas == '64-17-5' and k == '발암성'):
							cmr_map[k].append(v)
							if v not in merged[k]:
								merged[k].append(v)
					else:
						if v not in merged[k]:
							merged[k].append(v)

			if inhalation_entries:
				result['급성 독성(흡입)'] = '\n'.join(inhalation_entries)

			for k, v_list in merged.items():
				if cas == '64-17-5' and k == '발암성':
					continue
				result[k] = '\n'.join(v_list)
				if k not in HAZARD_ORDER:
					unknown_cols.add(k)

			cmr_grades = []
			for values in cmr_map.values():
				for v in values:
					cmr_grades.extend(extract_cmr_grades(v))
			highest_grade = get_highest_cmr_grade(cmr_grades)
			if highest_grade:
				result['CMR'] = highest_grade

			for agg_col, source_keys in AGGREGATE_GROUPS:
				grade = compute_aggregate_grade(result, source_keys)
				if grade:
					result[agg_col] = grade

		# 상세 API는 동시에 조회
		t_twa = asyncio.create_task(_query_twa_stel_async(session, semaphore, service_key, chem_id))
		t_vp = asyncio.create_task(_query_vapor_pressure_async(session, semaphore, service_key, chem_id))
		t_rev = asyncio.create_task(_query_revision_date_async(session, semaphore, service_key, chem_id))
		t_15 = asyncio.create_task(_query_detail15_async(session, semaphore, service_key, chem_id))
		t_11 = asyncio.create_task(_query_detail11_async(session, semaphore, service_key, chem_id))

		(twa_stel, vapor_pressure, revision_date, detail15_map, stot_pair) = await asyncio.gather(
			t_twa, t_vp, t_rev, t_15, t_11
		)

		result['TWA'], result['STEL'] = twa_stel
		result['증기압'] = vapor_pressure
		result['개정일'] = revision_date

		if detail15_map['관리대상유해물질']:
			result['관리대상유해물질'] = '▣'
		if detail15_map['특별관리물질']:
			result['특별관리물질'] = '▣'
		if detail15_map['특수건강진단대상물질']:
			result['특수건강진단대상물질'] = '▣'
		if detail15_map['인체급성유해성물질']:
			result['인체급성유해성물질'] = '▣'
		if detail15_map['인체만성유해성물질']:
			result['인체만성유해성물질'] = '▣'
		if detail15_map['생태유해성물질']:
			result['생태유해성물질'] = '▣'
		if detail15_map['작업환경측정대상물질']:
			result['작업환경측정대상물질'] = '▣'
		if detail15_map['노출기준설정물질']:
			result['노출기준설정물질'] = '▣'
		if detail15_map['허용기준설정물질']:
			result['허용기준설정물질'] = '▣'
		if detail15_map['금지물질']:
			result['금지물질'] = '▣'
		if detail15_map['제한물질']:
			result['제한물질'] = '▣'
		if detail15_map['허가물질']:
			result['허가물질'] = '▣'
		if detail15_map['사고대비물질']:
			result['사고대비물질'] = '▣'
		if detail15_map['중점관리물질']:
			result['중점관리물질'] = '▣'
		if detail15_map['위험물']:
			result['위험물'] = '▣'
		if detail15_map['독성가스']:
			result['독성가스'] = '▣'

		stot_single, stot_repeated = stot_pair
		if stot_single:
			result['[11번] 특정표적장기 독성(1회 노출)'] = stot_single
		if stot_repeated:
			result['[11번] 특정표적장기 독성(반복 노출)'] = stot_repeated

	except Exception as e:
		result['결과없음'] = f'조회 오류: {str(e)}'

	return idx, result, unknown_cols


async def _query_cas_info_async(data_rows, service_key, progress_callback=None):
	unknown_columns = set()
	processed_data = []
	total = len(data_rows)
	completed = 0

	semaphore = asyncio.Semaphore(ASYNC_MAX_CONCURRENCY)
	timeout = aiohttp.ClientTimeout(total=20, connect=5, sock_read=15)
	connector = aiohttp.TCPConnector(limit=ASYNC_MAX_CONCURRENCY)

	async with aiohttp.ClientSession(timeout=timeout, connector=connector) as session:
		tasks = [
			asyncio.create_task(process_single_chemical_async(idx, row, session, service_key, semaphore))
			for idx, row in enumerate(data_rows.itertuples(index=False), start=1)
		]

		for task in asyncio.as_completed(tasks):
			idx, res, unk_cols = await task
			processed_data.append((idx, res))
			unknown_columns.update(unk_cols)
			completed += 1
			if progress_callback is not None:
				progress_callback(completed, total)

	processed_data.sort(key=lambda x: x[0])
	results = [x[1] for x in processed_data]
	return pd.DataFrame(results), sorted(list(unknown_columns))


def _query_cas_info_threaded(data_rows, service_key, progress_callback=None):
	"""기존 스레드 기반 조회 경로(폴백)"""
	unknown_columns = set()
	processed_data = []
	total = len(data_rows)
	completed = 0

	with requests.Session() as session:
		with concurrent.futures.ThreadPoolExecutor(max_workers=10) as executor:
			future_to_idx = {
				executor.submit(process_single_chemical, idx, row, session, service_key): idx
				for idx, row in enumerate(data_rows.itertuples(index=False), start=1)
			}

			for future in concurrent.futures.as_completed(future_to_idx):
				idx, res, unk_cols = future.result()
				processed_data.append((idx, res))
				unknown_columns.update(unk_cols)
				completed += 1
				if progress_callback is not None:
					progress_callback(completed, total)

	processed_data.sort(key=lambda x: x[0])
	results = [x[1] for x in processed_data]
	return pd.DataFrame(results), sorted(list(unknown_columns))


def query_cas_info(data_rows, service_key, progress_callback=None):
	"""여러 물질 조회를 비동기로 우선 시도하고, 불가하면 스레드 경로로 폴백한다."""
	if aiohttp is None:
		return _query_cas_info_threaded(data_rows, service_key, progress_callback=progress_callback)

	try:
		asyncio.get_running_loop()
		# 이미 러닝 루프가 있으면 안전하게 기존 경로를 사용
		return _query_cas_info_threaded(data_rows, service_key, progress_callback=progress_callback)
	except RuntimeError:
		# 러닝 루프가 없으면 비동기 경로 사용
		try:
			return asyncio.run(_query_cas_info_async(data_rows, service_key, progress_callback=progress_callback))
		except Exception:
			return _query_cas_info_threaded(data_rows, service_key, progress_callback=progress_callback)


def _write_excel_results(wb, ws, hazard_df):
	"""KOSHA 조회 결과를 기존 워크북에 써 넣고 표2~표4를 만든다."""
	col_name_to_idx = {cell.value: idx for idx, cell in enumerate(ws[1], start=1)}

	# 헤더 스타일 보정(기존 makeResult 동작 유지)
	for row in ws['A1:BY1']:
		for cell in row:
			cell.fill = fill_header
			cell.font = default_font
			if cell.value in ['발암성', '생식독성', 'CMR', '급성 독성', '피부/눈 자극성', '특정표적장기 독성', '수생환경 유해성', '인화성', '연간사용·판매량']:
				cell.font = bold_font

	# 본문 값 채우기
	for r_idx, row in hazard_df.iterrows():
		excel_row = r_idx + 2
		for col_name in HAZARD_ORDER:
			if col_name in col_name_to_idx:
				col_idx = col_name_to_idx[col_name]
				cell = ws.cell(row=excel_row, column=col_idx)
				cell.value = row.get(col_name, '')

				col_letter = get_column_letter(col_idx)
				if col_letter in ['A', 'B', 'C']:
					cell.fill = fill_label
				elif col_letter in ['BK', 'BL', 'BM', 'BN', 'BO', 'BP', 'BQ', 'BR', 'BS', 'BT', 'BU', 'BV', 'BW']:
					cell.fill = fill_data2
				else:
					cell.fill = fill_data1

				if col_name in [
					'#', 'CAS No.', '결과없음', '개정일',
					'관리대상유해물질', '특별관리물질', '특수건강진단대상물질', '인체급성유해성물질', '인체만성유해성물질',
					'생태유해성물질', '작업환경측정대상물질', '노출기준설정물질', '허용기준설정물질', '금지물질', '제한물질',
					'허가물질', '사고대비물질', '중점관리물질', '위험물', '독성가스',
				]:
					cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
				else:
					if col_name in ['[11번] 특정표적장기 독성(1회 노출)', '[11번] 특정표적장기 독성(반복 노출)']:
						cell.alignment = Alignment(vertical='top', wrap_text=False)
					else:
						cell.alignment = Alignment(vertical='center', wrap_text=True)

	_build_summary_tables(ws, hazard_df)


def _build_summary_tables(ws, hazard_df):
	"""표2(유해성), 표3(규제물질), 표4(연간입고/사용량)을 생성한다."""
	thin_border = Border(
		left=Side(style='thin'), right=Side(style='thin'),
		top=Side(style='thin'), bottom=Side(style='thin'),
	)

	# -------- 표2 --------
	start_idx = HAZARD_ORDER.index('발암성')
	end_idx = HAZARD_ORDER.index('금속부식성 물질') + 1
	hazard_cols = HAZARD_ORDER[start_idx:end_idx]
	hazard_start_col = 4

	start_row = 2
	end_row = start_row + len(hazard_df) - 1
	summary_start_row = end_row + 2

	excluded_result_values = {'공단 MSDS 없음', '영업비밀', '심의중'}
	analyzed_count = sum(
		1
		for r in range(start_row, end_row + 1)
		if str(ws.cell(row=r, column=HAZARD_ORDER.index('결과없음') + 1).value or '').strip() not in excluded_result_values
	)

	ws[f'D{summary_start_row}'] = '유해성'
	ws[f'D{summary_start_row}'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
	ws[f'D{summary_start_row}'].font = default_font
	ws[f'D{summary_start_row}'].border = thin_border
	ws[f'D{summary_start_row}'].fill = fill_header

	for idx, col_name in enumerate(hazard_cols):
		col_letter = get_column_letter(hazard_start_col + idx + 1)
		cell = ws[f'{col_letter}{summary_start_row}']
		cell.value = col_name
		cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
		cell.font = bold_font if col_name in ['발암성', '생식독성', 'CMR', '급성 독성', '피부/눈 자극성', '특정표적장기 독성', '수생환경 유해성', '인화성'] else default_font
		cell.border = thin_border
		cell.fill = fill_header

	row_labels = ['구분1', '1A', '1B', '구분2', '구분3', '구분4', '기타구분', '유해물질수', '분석물질수', '유해물질비율']
	cmr_agg_cols = {'CMR'}
	general_agg_cols = {agg_col for agg_col, _ in AGGREGATE_GROUPS}
	cmr_source_cols = {'발암성', '생식독성', '생식세포 변이원성'}
	summary_data = []

	for hazard in hazard_cols:
		col_idx = HAZARD_ORDER.index(hazard) + 1
		count_map = {label: 0 for label in row_labels[:-3]}

		for r in range(start_row, end_row + 1):
			val = ws.cell(row=r, column=col_idx).value
			if val is None or pd.isna(val):
				continue
			if str(val).strip() == '':
				continue
			val_str = str(val).strip()
			if val_str.lower() in {'nan', 'none', 'null'}:
				continue

			if hazard in cmr_agg_cols:
				label_map = {'1A': '1A', '1B': '1B', '2': '구분2'}
				count_map[label_map.get(val_str, '기타구분')] += 1
			elif hazard in general_agg_cols:
				label_map = {'1': '구분1', '2': '구분2', '3': '구분3', '4': '구분4'}
				count_map[label_map.get(val_str, '기타구분')] += 1
			elif hazard in cmr_source_cols:
				grades = extract_cmr_grades(val_str)
				most_severe = get_highest_cmr_grade(grades)
				label_map = {'1A': '1A', '1B': '1B', '2': '구분2'}
				if most_severe:
					count_map[label_map.get(most_severe, '기타구분')] += 1
			else:
				grades = extract_grades(val_str)
				most_severe = get_highest_grade(grades)
				label_map = {'1': '구분1', '2': '구분2', '3': '구분3', '4': '구분4'}
				if most_severe:
					count_map[label_map.get(most_severe, '기타구분')] += 1
				else:
					count_map['기타구분'] += 1

		count_map['유해물질수'] = sum(count_map[label] for label in row_labels[:7])
		count_map['분석물질수'] = analyzed_count
		count_map['유해물질비율'] = f"{round((count_map['유해물질수'] / analyzed_count) * 100)}%" if analyzed_count else '0%'

		for i, label in enumerate(row_labels):
			if len(summary_data) <= i:
				summary_data.append([])
			summary_data[i].append(count_map[label])

	for row_offset, (label, row_values) in enumerate(zip(row_labels, summary_data), start=1):
		row_num = summary_start_row + row_offset
		label_cell = ws.cell(row=row_num, column=hazard_start_col)
		label_cell.value = label
		label_cell.alignment = Alignment(horizontal='center', vertical='center')
		label_cell.font = default_font
		label_cell.border = thin_border
		label_cell.fill = fill_label

		for col_offset, value in enumerate(row_values):
			col_num = hazard_start_col + col_offset + 1
			cell = ws.cell(row=row_num, column=col_num)
			if label in ('유해물질수', '유해물질비율'):
				cell.value = value
			else:
				cell.value = value if value != 0 else None
			cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
			cell.font = default_font
			cell.border = thin_border

	# -------- 표3 --------
	summary_titles = [
		'관리대상유해물질', '특별관리물질', '작업환경측정대상물질', '특수건강진단대상물질',
		'노출기준설정물질', '허용기준설정물질', '금지물질', '제한물질',
		'인체급성유해성물질', '인체만성유해성물질', '생태유해성물질',
		'허가물질', '사고대비물질', '중점관리물질', '위험물', '독성가스',
		'인체등유해성물질',
		'제한물질2', '금지물질2', '허가물질2', '사고대비물질2',
		'중점관리물질2', '금지·허가물질2', '노출·허용기준물질2', '직업환경측정물질등2', '위험물2', '독성가스2',
	]
	summary_start_col = 47
	table3_start_row = end_row + 2

	analyzed_count2 = 0
	col_result_idx = HAZARD_ORDER.index('결과없음') + 1
	for r in range(start_row, end_row + 1):
		val = str(ws.cell(row=r, column=col_result_idx).value or '').strip()
		if val not in excluded_result_values:
			analyzed_count2 += 1

	c = ws.cell(row=table3_start_row, column=summary_start_col - 1)
	c.value = '규제물질'
	c.alignment = Alignment(horizontal='center', vertical='center')
	c.font = default_font
	c.border = thin_border
	c.fill = fill_header

	c = ws.cell(row=table3_start_row + 1, column=summary_start_col - 1)
	c.value = '물질 수'
	c.alignment = Alignment(horizontal='center', vertical='center')
	c.font = default_font
	c.border = thin_border
	c.fill = fill_label

	c = ws.cell(row=table3_start_row + 2, column=summary_start_col - 1)
	c.value = '물질 비율'
	c.alignment = Alignment(horizontal='center', vertical='center')
	c.font = default_font
	c.border = thin_border
	c.fill = fill_label

	for idx, col_name in enumerate(summary_titles):
		cell = ws.cell(row=table3_start_row, column=summary_start_col + idx, value=col_name)
		cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
		cell.font = default_font
		cell.border = thin_border
		cell.fill = fill_header

	for idx in range(len(summary_titles)):
		col_idx = summary_start_col + idx
		count = sum(1 for r in range(start_row, end_row + 1) if str(ws.cell(row=r, column=col_idx).value).strip() == '▣')

		cell = ws.cell(row=table3_start_row + 1, column=col_idx, value=count)
		cell.alignment = Alignment(horizontal='center', vertical='center')
		cell.font = default_font
		cell.border = thin_border

		ratio = f'{round((count / analyzed_count2) * 100)}%' if analyzed_count2 else '0%'
		cell = ws.cell(row=table3_start_row + 2, column=col_idx, value=ratio)
		cell.alignment = Alignment(horizontal='center', vertical='center')
		cell.font = default_font
		cell.border = thin_border

	# -------- 표4 --------
	def _normalize_header(value):
		if value is None:
			return ''
		return str(value).strip().replace(' ', '')

	col_idx_in = None
	col_idx_use = None
	for col in range(1, ws.max_column + 1):
		header = ws.cell(row=1, column=col).value
		norm = _normalize_header(header)
		if norm == '연간입고량':
			col_idx_in = col
		elif norm in {'연간사용·판매량', '연간사용판매량'}:
			col_idx_use = col

	if col_idx_in is None or col_idx_use is None:
		return

	table3_end_row = table3_start_row + 2
	table4_start_row = table3_end_row + 2
	table4_start_col = 73
	headers4 = ['중량(톤/년) 또는 부피단위(㎥/년)', '연간입고량', '연간사용·판매량']

	for idx, header in enumerate(headers4):
		cell = ws.cell(row=table4_start_row, column=table4_start_col + idx)
		cell.value = header
		cell.font = default_font
		cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
		cell.border = thin_border
		cell.fill = fill_header

	usage_map2 = {
		'1': '0.1미만', '2': '0.1~0.5', '3': '0.5~1.0', '4': '1~2.5', '5': '2.5~5.0',
		'6': '5~20', '7': '20~200', '8': '200~1,000', '9': '1,000~5,000', '10': '5,000이상',
		1: '0.1미만', 2: '0.1~0.5', 3: '0.5~1.0', 4: '1~2.5', 5: '2.5~5.0',
		6: '5~20', 7: '20~200', 8: '200~1,000', 9: '1,000~5,000', 10: '5,000이상',
		1.0: '0.1미만', 2.0: '0.1~0.5', 3.0: '0.5~1.0', 4.0: '1~2.5', 5.0: '2.5~5.0',
		6.0: '5~20', 7.0: '20~200', 8.0: '200~1,000', 9.0: '1,000~5,000', 10.0: '5,000이상',
	}
	usage_descriptions = ['0.1미만', '0.1~0.5', '0.5~1.0', '1~2.5', '2.5~5.0', '5~20', '20~200', '200~1,000', '1,000~5,000', '5,000이상']

	def _norm(v):
		if v is None or isinstance(v, bool):
			return None
		if isinstance(v, (int, float)):
			if v in usage_map2:
				return usage_map2[v]
			try:
				return usage_map2.get(int(v), None)
			except Exception:
				return None

		raw = str(v).strip()
		if raw == '':
			return None

		norm = raw.replace('\u3000', ' ').replace('，', ',').replace('–', '-').replace('\u2013', '-')
		norm = norm.replace(' ', '')

		for desc in usage_descriptions:
			if desc.replace(',', '') in norm or desc.replace('，', '') in norm:
				return desc

		if '미만' in raw:
			return '0.1미만'
		if '이상' in raw:
			return '5,000이상'

		m = re.search(r'\b([1-9]|10)\b', raw)
		if m:
			try:
				return usage_map2.get(int(m.group(1)))
			except Exception:
				pass

		m2 = re.match(r'^\s*(\d+)(?:\.0+)?\s*$', raw)
		if m2:
			try:
				return usage_map2.get(int(m2.group(1)))
			except Exception:
				pass

		digits = re.sub(r'[^0-9]', '', raw)
		if digits:
			try:
				di = int(digits)
				if 1 <= di <= 10:
					return usage_map2.get(di)
			except Exception:
				pass

		return None

	incoming_counter = Counter()
	usage_counter = Counter()

	for r in range(start_row, end_row + 1):
		desc_in = _norm(ws.cell(row=r, column=col_idx_in).value)
		desc_use = _norm(ws.cell(row=r, column=col_idx_use).value)

		if desc_in:
			ws.cell(row=r, column=col_idx_in).value = desc_in
			incoming_counter[desc_in] += 1
		if desc_use:
			ws.cell(row=r, column=col_idx_use).value = desc_use
			usage_counter[desc_use] += 1

	for i, desc in enumerate(usage_descriptions):
		row = table4_start_row + 1 + i

		c_desc = ws.cell(row=row, column=table4_start_col)
		c_desc.value = desc
		c_desc.font = default_font
		c_desc.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
		c_desc.border = thin_border
		c_desc.fill = fill_label

		v_in = incoming_counter.get(desc, 0)
		c_in = ws.cell(row=row, column=table4_start_col + 1)
		c_in.value = v_in if v_in != 0 else None
		c_in.font = default_font
		c_in.alignment = Alignment(horizontal='center', vertical='center')
		c_in.border = thin_border

		v_use = usage_counter.get(desc, 0)
		c_use = ws.cell(row=row, column=table4_start_col + 2)
		c_use.value = v_use if v_use != 0 else None
		c_use.font = default_font
		c_use.alignment = Alignment(horizontal='center', vertical='center')
		c_use.border = thin_border


def _extract_data_rows_from_workbook(ws):
	"""A/B/C열에서 조회 대상 데이터프레임(#, 물질명칭, CAS No.)을 만든다."""
	data = []
	row_no = 2
	while True:
		idx_val = ws.cell(row=row_no, column=1).value
		name_val = ws.cell(row=row_no, column=2).value
		cas_val = ws.cell(row=row_no, column=3).value

		# 세 칸 모두 비었으면 종료한다.
		if idx_val is None and name_val is None and cas_val is None:
			break

		data.append({'#': idx_val, '물질명칭': name_val, 'CAS No.': cas_val})
		row_no += 1

	df = pd.DataFrame(data)
	if df.empty:
		return df

	# 기존 makeResult 규칙: CAS No. 문자열화 후 좌측 0 제거
	df['CAS No.'] = df['CAS No.'].astype(str).str.lstrip('0')
	return df[['#', '물질명칭', 'CAS No.']].copy()


def fill_kosha_into_workbook_bytes(base_excel_bytes, service_key, progress_callback=None):
	"""ICIS 기본 엑셀 바이트를 받아 KOSHA 반영 + 표2~표4 생성 후 바이트 반환."""
	wb = load_workbook(io.BytesIO(base_excel_bytes))
	ws = wb.active

	data_rows = _extract_data_rows_from_workbook(ws)
	if data_rows.empty:
		# 물질이 없으면 원본 그대로 반환
		out = io.BytesIO()
		wb.save(out)
		out.seek(0)
		return out.getvalue()

	hazard_df, _ = query_cas_info(data_rows, service_key, progress_callback=progress_callback)
	_write_excel_results(wb, ws, hazard_df)

	out = io.BytesIO()
	wb.save(out)
	out.seek(0)
	return out.getvalue()


def create_final_excel_bytes_for_company(session, keyword, company_item, progress_callback=None):
	"""
	통합 핵심 함수
	1) ICIS 템플릿 생성
	2) 같은 메모리에서 즉시 KOSHA 채움
	3) 표2~표4까지 완성된 최종 바이트 반환
	"""
	company_name_full, base_excel_bytes = create_excel_bytes_for_company(
		session=session,
		company_name=keyword,
		item=company_item,
		search_year=None,
	)
	final_excel_bytes = fill_kosha_into_workbook_bytes(
		base_excel_bytes,
		SERVICE_KEY,
		progress_callback=progress_callback,
	)
	return company_name_full, final_excel_bytes


# =========================================================
# 3) Streamlit UI (단일 화면)
# =========================================================
def toggle_select_all():
	"""전체 선택 체크박스 상태를 각 행 체크박스에 반영한다."""
	val = st.session_state['select_all_main']
	st.session_state['select_all_state'] = val
	page_indices = st.session_state.get('current_page_indices', [])
	for i in page_indices:
		st.session_state[f'chk_{i}'] = val


def check_individual_toggle():
	"""개별 체크박스 변화를 감지해 전체선택 체크 상태를 동기화한다."""
	page_indices = st.session_state.get('current_page_indices', [])
	all_checked = bool(page_indices) and all(st.session_state.get(f'chk_{i}', False) for i in page_indices)
	st.session_state['select_all_main'] = all_checked
	st.session_state['select_all_state'] = all_checked


def _init_session_state():
	"""화면에서 쓰는 상태값을 한 번에 초기화한다."""
	defaults = {
		'search_results': [],
		'files': {},
		'select_all_state': False,
		'select_all_main': False,
		'current_page': 1,
		'current_page_indices': [],
		'page_size': 10,
		'keywords': '',
		'search_success_message': None,
		'last_search_keywords': None,
		'file_generation_started': False,
		'active_generation_indices': [],
		'file_generation_completed': False,
		'page_downloads': {},
		'reset_requested': False,
		'download_bytes_data': None,
		'download_filename': '',
		'download_mime_type': 'application/zip',
		'download_completed': False,
	}
	for key, value in defaults.items():
		if key not in st.session_state:
			st.session_state[key] = value


def _render_status_cell(idx, text_placeholder, bar_placeholder):
	"""상태 셀에 업체별 진행 상태(텍스트+진행바)를 렌더링한다."""
	status = st.session_state.get(f'status_{idx}', '')
	progress_value = int(st.session_state.get(f'progress_{idx}', 0) or 0)
	progress_value = max(0, min(progress_value, 100))

	if status == '✅':
		text_placeholder.markdown('<div style="text-align: center; font-size: 12px;">완료 100%</div>', unsafe_allow_html=True)
		bar_placeholder.progress(1.0)
	elif status == '❌':
		text_placeholder.markdown(f'<div style="text-align: center; font-size: 12px; color: #c5221f;">오류 {progress_value}%</div>', unsafe_allow_html=True)
		bar_placeholder.progress(progress_value / 100 if progress_value else 0.0)
	elif status == '처리중':
		text_placeholder.markdown(f'<div style="text-align: center; font-size: 12px;">진행중 {progress_value}%</div>', unsafe_allow_html=True)
		bar_placeholder.progress(progress_value / 100)
	else:
		text_placeholder.markdown('<div style="text-align: center; font-size: 12px; color: #777;">대기 0%</div>', unsafe_allow_html=True)
		bar_placeholder.progress(0.0)


def _clear_search_runtime_state(clear_keyword=False):
	"""검색/다운로드 관련 세션 상태를 정리한다."""
	prev_results = st.session_state.get('search_results', [])
	for i in range(len(prev_results)):
		st.session_state[f'chk_{i}'] = False
		st.session_state[f'status_{i}'] = ''
		st.session_state[f'progress_{i}'] = 0
		st.session_state[f'filename_{i}'] = ''

	if clear_keyword:
		st.session_state['keyword_input'] = ''

	st.session_state['search_results'] = []
	st.session_state['current_page'] = 1
	st.session_state['current_page_indices'] = []
	st.session_state['search_success_message'] = None
	st.session_state['last_search_keywords'] = None
	st.session_state['select_all_state'] = False
	st.session_state['select_all_main'] = False
	st.session_state['files'] = {}
	st.session_state['page_downloads'] = {}
	st.session_state['file_generation_started'] = False
	st.session_state['active_generation_indices'] = []
	st.session_state['file_generation_completed'] = False
	st.session_state['download_bytes_data'] = None
	st.session_state['download_filename'] = ''
	st.session_state['download_mime_type'] = 'application/zip'
	st.session_state['download_completed'] = False


def main_ui(tab_mode=False):
	"""검색 UI는 createForm과 동일, 생성 동작만 통합 파이프라인으로 연결한다."""
	_init_session_state()

	# 검색 컨테이너 바로 위에 제목/버전을 복원한다.
	col_tab, col_ver = st.columns([0.85, 0.15])
	with col_tab:
		st.markdown('## 화학물질 정보 수집 시스템')
	with col_ver:
		st.markdown(
			"""<div style="text-align: right; color: #999; font-size: 15px; margin-top: 10px;">
			v2.260705
			</div>""",
			unsafe_allow_html=True,
		)

	# createForm과 동일한 검색 카드/테이블 스타일
	st.markdown("""
	<style>
	/* 기본 공통 버튼 스타일 */
	button {
		background-color: #f0f5f9 !important;
		color: #555 !important;
		border: 1px solid #d0d7de !important;
		font-weight: 500;
	}
	button:hover {
		background-color: #e5ebf2 !important;
		border-color: #b1bac4 !important;
	}

	/* st.form 스타일 자체를 테두리 둥근 카드로 리디자인 */
	div[data-testid="stForm"] {
		border: 1px solid #e6e8eb !important;
		border-radius: 12px !important;
		padding: 20px !important;
		background-color: #ffffff !important;
		box-shadow: 0 4px 6px rgba(0,0,0,0.02) !important;
	}

	/* 입력창 실제 입력 필드의 배경을 흰색으로 지정하고 플레이스홀더 글자크기(14px)와 완벽 통일 */
	div[data-testid="stTextInput"] input {
		background-color: #ffffff !important;
		color: #31333F !important;
		font-size: 14px !important;
		font-family: inherit !important;
		border: 1px solid #d0d7de !important;
		border-radius: 6px !important;
	}
	/* 포커스 상태에서도 흰색 배경 유지 */
	div[data-testid="stTextInput"] input:focus {
		background-color: #ffffff !important;
		color: #31333F !important;
		border: 1px solid #d0d7de !important;
		outline: none !important;
		box-shadow: none !important;
	}
	/* 자동 완성(autofill) 상태에서도 흰색 배경 유지 */
	div[data-testid="stTextInput"] input:-webkit-autofill {
		-webkit-box-shadow: 0 0 0 1000px white inset !important;
		-webkit-text-fill-color: #31333F !important;
		background-color: #ffffff !important;
	}
	div[data-testid="stTextInput"] input:-webkit-autofill:focus {
		-webkit-box-shadow: 0 0 0 1000px white inset !important;
		-webkit-text-fill-color: #31333F !important;
		background-color: #ffffff !important;
	}
	div[data-testid="stTextInput"] input::placeholder {
		font-size: 14px !important;
		font-family: inherit !important;
		color: #8c959f !important;
	}

	/* 테이블용 st.columns의 모든 행 내부 요소 세로 중간 정렬 고정 */
	div.table-row-container div[data-testid="column"] {
		display: flex !important;
		align-items: center !important;
		min-height: 40px !important;
	}

	/* 텍스트 하단 마진 제거하여 체크박스와 가로 수평 완벽 동기화 */
	div.table-row-container div[data-testid="column"] p,
	div.table-row-container div[data-testid="column"] div,
	div.table-row-container div[data-testid="column"] span {
		margin: 0 !important;
		padding: 0 !important;
		line-height: 1.2 !important;
	}

	/* 테이블 내부의 텍스트만 2.5px 아래로 수평 수선 */
	div.table-row-container div[data-testid="column"]:not(:nth-of-type(1)) p,
	div.table-row-container div[data-testid="column"]:not(:nth-of-type(1)) div,
	div.table-row-container div[data-testid="column"]:not(:nth-of-type(1)) span {
		transform: translateY(2.5px) !important;
	}

	/* 체크박스 마진 제거 */
	div[data-testid="stCheckbox"] {
		margin: 0 !important;
		padding: 0 !important;
		display: flex !important;
		align-items: center !important;
		justify-content: center !important;
	}

	/* 체크박스 하위 label 세로 정밀 1:1 중간 정렬 */
	div[data-testid="stCheckbox"] > label {
		display: flex !important;
		align-items: center !important;
		justify-content: center !important;
		margin: 0 !important;
		padding: 0 !important;
		min-height: unset !important;
	}

	/* No, 지역, 상태 열(1, 2, 4, 5번째 열) 내용 가운데 가로 정렬 */
	div.table-row-container div[data-testid="column"]:nth-of-type(1),
	div.table-row-container div[data-testid="column"]:nth-of-type(2),
	div.table-row-container div[data-testid="column"]:nth-of-type(4),
	div.table-row-container div[data-testid="column"]:nth-of-type(5) {
		justify-content: center !important;
	}

	/* 업체명 및 파일명 열(3, 6번째 열) 내용 왼쪽 정렬 유지 */
	div.table-row-container div[data-testid="column"]:nth-of-type(3),
	div.table-row-container div[data-testid="column"]:nth-of-type(6) {
		justify-content: flex-start !important;
	}

	/* 테이블 첫 번째 컬럼(체크박스 영역) 디자인 */
	div.table-row-container div[data-testid="column"]:nth-of-type(1) {
		background-color: #f0f0f0;
		padding: 4px;
		border-radius: 4px;
		display: flex;
		justify-content: center;
		align-items: center;
	}

	/* 💥 [하단 카드 정교화] 가로/세로 오차 없는 정밀 높이 바인딩 및 둥근 모서리 마감 */
	.bottom-card-anchor + div[data-testid="stContainer"] {
		border-radius: 12px !important;
		padding: 16px 20px !important;
		background-color: #ffffff !important;
		border: 1px solid #e6e8eb !important;
		box-shadow: 0 4px 6px rgba(0,0,0,0.02) !important;
	}

	/* 하단 카드 내부의 columns 가로 정렬 및 여백 조율 */
	.bottom-card-anchor + div[data-testid="stContainer"] div[data-testid="stHorizontalBlock"] {
		display: flex !important;
		align-items: stretch !important;
		gap: 16px !important;
		height: 84px !important;
	}

	/* [하단 카드 전용 column 정밀 override] */
	.bottom-card-anchor + div[data-testid="stContainer"] div[data-testid="column"] {
		display: flex !important;
		align-items: stretch !important;
		min-height: 84px !important;
		height: 84px !important;
	}

	/* [높이 일치 마감] 왼쪽 column의 모든 래퍼와 빈 공간의 높이를 84px로 강제 통일 */
	.bottom-card-anchor + div[data-testid="stContainer"] div[data-testid="column"]:nth-of-type(1) {
		display: flex !important;
		align-items: stretch !important;
		justify-content: flex-start !important;
		width: 85% !important;
		height: 84px !important;
		min-height: 84px !important;
	}

	.bottom-card-anchor + div[data-testid="stContainer"] div[data-testid="column"]:nth-of-type(1) > div[data-testid="stVerticalBlock"],
	.bottom-card-anchor + div[data-testid="stContainer"] div[data-testid="column"]:nth-of-type(1) div.element-container,
	.bottom-card-anchor + div[data-testid="stContainer"] div[data-testid="column"]:nth-of-type(1) div.stMarkdown,
	.bottom-card-anchor + div[data-testid="stContainer"] div[data-testid="column"]:nth-of-type(1) div.stMarkdown > div {
		height: 100% !important;
		margin: 0 !important;
		padding: 0 !important;
		display: flex !important;
		align-items: stretch !important;
		width: 100% !important;
	}

	/* [높이 일치 마감] 오른쪽 버튼 column 내부 stVerticalBlock에 flex 정렬 부여하여 84px 완벽 실장 */
	.bottom-card-anchor + div[data-testid="stContainer"] div[data-testid="column"]:nth-of-type(2) {
		display: flex !important;
		flex-direction: column !important;
		justify-content: space-between !important;
		align-items: stretch !important;
		width: 15% !important;
		height: 84px !important;
		min-height: 84px !important;
	}

	.bottom-card-anchor + div[data-testid="stContainer"] div[data-testid="column"]:nth-of-type(2) > div[data-testid="stVerticalBlock"] {
		display: flex !important;
		flex-direction: column !important;
		justify-content: space-between !important;
		height: 84px !important;
		gap: 8px !important;
	}

	.bottom-card-anchor + div[data-testid="stContainer"] div[data-testid="column"]:nth-of-type(2) div.element-container {
		margin: 0 !important;
		padding: 0 !important;
		height: 38px !important;
	}

	/* 하단 카드 전용 수직 버튼 세트 규격 강제 통일 */
	.bottom-card-anchor + div[data-testid="stContainer"] button {
		height: 38px !important;
		min-height: 38px !important;
		margin: 0 !important;
		padding: 0px 16px !important;
		line-height: 38px !important;
		display: flex !important;
		align-items: center !important;
		justify-content: center !important;
		width: 100% !important;
		box-sizing: border-box !important;
	}

	/* 검색 폼 내부 라벨 수평 정렬용 텍스트 보정 */
	p.search-label-txt {
		font-size: 16px;
		font-weight: bold;
		margin: 0 !important;
		padding: 0 !important;
		white-space: nowrap;
		transform: translateY(8px);
	}

	/* 페이지 네비게이션 전용 버튼 스타일 */
	button[kind="tertiary"] {
		background: transparent !important;
		background-color: transparent !important;
		border: none !important;
		outline: none !important;
		box-shadow: none !important;
		color: #666 !important;
		padding: 0 !important;
		margin: 0 !important;
		font-size: 14px !important;
		font-weight: 400 !important;
		line-height: 1.3 !important;
	}
	button[kind="tertiary"]:hover,
	button[kind="tertiary"]:focus,
	button[kind="tertiary"]:focus-visible,
	button[kind="tertiary"]:active,
	button[kind="tertiary"]:disabled {
		background: transparent !important;
		background-color: transparent !important;
		border: none !important;
		outline: none !important;
		box-shadow: none !important;
	}
	/* 선택된 페이지 번호(비활성 tertiary 버튼)는 굵게 */
	.page-nav-anchor + div button[kind="tertiary"]:disabled {
		opacity: 1 !important;
		color: #666 !important;
		font-size: 14px !important;
		line-height: 1.3 !important;
		font-family: inherit !important;
		font-weight: 800 !important;
	}
	.page-nav-anchor + div button[kind="tertiary"]:not(:disabled) {
		font-weight: 400 !important;
	}
	.page-nav-anchor + div button[kind="tertiary"]:not(:disabled) * {
		font-weight: 400 !important;
	}
	.page-nav-anchor + div button[kind="tertiary"]:disabled * {
		font-size: 14px !important;
		line-height: 1.3 !important;
		font-family: inherit !important;
		font-weight: 800 !important;
	}
	.page-nav-anchor + div[data-testid="stHorizontalBlock"] div[data-testid="stButton"] {
		display: flex !important;
		justify-content: center !important;
	}
	.page-nav-anchor + div[data-testid="stHorizontalBlock"] button {
		background: transparent !important;
		background-color: transparent !important;
		color: #666 !important;
		border: none !important;
		outline: none !important;
		box-shadow: none !important;
		-webkit-appearance: none !important;
		appearance: none !important;
		height: auto !important;
		min-height: auto !important;
		padding: 0 !important;
		margin: 0 !important;
		font-size: 14px !important;
		font-weight: 400 !important;
		line-height: 1.3 !important;
	}
	.page-nav-anchor + div[data-testid="stHorizontalBlock"] button[kind="primary"] {
		background: transparent !important;
		color: #111 !important;
		border: none !important;
	}
	.page-nav-anchor + div[data-testid="stHorizontalBlock"] button:hover {
		background: transparent !important;
		background-color: transparent !important;
		color: #111 !important;
		border: none !important;
		box-shadow: none !important;
	}
	.page-nav-anchor + div[data-testid="stHorizontalBlock"] button:focus,
	.page-nav-anchor + div[data-testid="stHorizontalBlock"] button:focus-visible,
	.page-nav-anchor + div[data-testid="stHorizontalBlock"] button:active,
	.page-nav-anchor + div[data-testid="stHorizontalBlock"] button:disabled {
		background: transparent !important;
		background-color: transparent !important;
		border: none !important;
		outline: none !important;
		box-shadow: none !important;
	}
	.page-nav-anchor + div[data-testid="stHorizontalBlock"] div[data-testid="column"] {
		padding-left: 6px !important;
		padding-right: 6px !important;
	}
	.page-current-text {
		text-align: center;
		font-size: 14px;
		font-weight: 700;
		color: #666;
		line-height: 1.3;
		margin: 0;
		padding: 0;
		font-family: inherit;
	}
	/* 꺽쇠(처음/이전/다음/마지막)만 살짝 크게 - aria-label 기반으로 강제 */
	.page-nav-anchor + div button[aria-label="«"],
	.page-nav-anchor + div button[aria-label="‹"],
	.page-nav-anchor + div button[aria-label="›"],
	.page-nav-anchor + div button[aria-label="»"] {
		font-size: 18px !important;
		line-height: 1.1 !important;
	}
	/* 페이징 번호 굵기 최종 강제: 선택(비활성)=bold, 나머지=normal */
	.page-nav-anchor + div[data-testid="stHorizontalBlock"] button:disabled,
	.page-nav-anchor + div[data-testid="stHorizontalBlock"] button[disabled] {
		font-weight: 800 !important;
	}
	.page-nav-anchor + div[data-testid="stHorizontalBlock"] button:disabled *,
	.page-nav-anchor + div[data-testid="stHorizontalBlock"] button[disabled] * {
		font-weight: 800 !important;
	}
	.page-nav-anchor + div[data-testid="stHorizontalBlock"] button:not(:disabled):not([disabled]) {
		font-weight: 400 !important;
	}
	.page-nav-anchor + div[data-testid="stHorizontalBlock"] button:not(:disabled):not([disabled]) * {
		font-weight: 400 !important;
	}
	</style>
	""", unsafe_allow_html=True)

	if st.session_state.get('reset_requested', False):
		_clear_search_runtime_state(clear_keyword=True)
		st.session_state['reset_requested'] = False

	# [상단 검색 카드] createForm과 동일 구조
	with st.form(key='search_form', clear_on_submit=False):
		col1, col2, col3 = st.columns([0.72, 0.14, 0.14])

		with col1:
			sub_col_lbl, sub_col_input = st.columns([0.08, 0.92])
			with sub_col_lbl:
				st.markdown('<p class="search-label-txt">업체 검색</p>', unsafe_allow_html=True)
			with sub_col_input:
				keyword_input = st.text_input(
					'업체 검색',
					value='',
					placeholder='예) 삼성전자, 엘지화학',
					key='keyword_input',
					label_visibility='collapsed',
				)
				search_msg_container = st.empty()

		search_button_pressed = False
		reset_button_pressed = False
		with col2:
			search_button_pressed = st.form_submit_button('🔍 검색', use_container_width=True)
		with col3:
			reset_button_pressed = st.form_submit_button('↺ 초기화', use_container_width=True)

	# 검색 초기화
	if reset_button_pressed:
		st.session_state['reset_requested'] = True
		st.rerun()

	# 검색 실행
	if search_button_pressed:
		_clear_search_runtime_state(clear_keyword=False)

		if not keyword_input.strip():
			search_msg_container.markdown('<p style="color: #e53e3e; font-size: 15px; font-weight: 500; margin-top: -10px; margin-bottom: 10px;">검색어를 입력하세요.</p>', unsafe_allow_html=True)
		else:
			keywords = [kw.strip() for kw in keyword_input.split(',') if kw.strip()]
			if not keywords:
				search_msg_container.markdown('<p style="color: #e53e3e; font-size: 15px; font-weight: 500; margin-top: -10px; margin-bottom: 10px;">유효한 검색어가 없습니다.</p>', unsafe_allow_html=True)
			else:
				st.session_state['keywords'] = ', '.join(keywords)
				try:
					search_results = search_companies(keywords)

					actual_results = [item for item in search_results if not item['company_name'].startswith('검색 결과 없음')]
					no_result_keywords = [item['search_keyword'] for item in search_results if item['company_name'].startswith('검색 결과 없음')]

					st.session_state['search_results'] = actual_results
					st.session_state['select_all_state'] = False
					st.session_state['select_all_main'] = False

					for i in range(len(actual_results)):
						st.session_state[f'chk_{i}'] = False
						st.session_state[f'status_{i}'] = ''
						st.session_state[f'progress_{i}'] = 0
						st.session_state[f'filename_{i}'] = ''

					if actual_results:
						st.session_state['search_success_message'] = f'{len(actual_results)}개의 업체를 찾았습니다.'
						search_msg_container.markdown(f'<p style="color: #e53e3e; font-size: 15px; font-weight: 500; margin-top: -10px; margin-bottom: 10px;">{st.session_state["search_success_message"]}</p>', unsafe_allow_html=True)
					elif no_result_keywords:
						st.session_state['search_success_message'] = '검색 결과가 없습니다'
						search_msg_container.markdown(f'<p style="color: #e53e3e; font-size: 15px; font-weight: 500; margin-top: -10px; margin-bottom: 10px;">{st.session_state["search_success_message"]}</p>', unsafe_allow_html=True)

					st.session_state['last_search_keywords'] = ', '.join(keywords)
				except Exception as exc:
					search_msg_container.markdown(f'<p style="color: #e53e3e; font-size: 15px; font-weight: 500; margin-top: -10px; margin-bottom: 10px;">검색 중 오류: {exc}</p>', unsafe_allow_html=True)

	search_results = st.session_state['search_results']

	if search_results:
		if st.session_state['search_success_message']:
			search_msg_container.markdown(f'<p style="color: #e53e3e; font-size: 15px; font-weight: 500; margin-top: -10px; margin-bottom: 10px;">{st.session_state["search_success_message"]}</p>', unsafe_allow_html=True)

		@st.fragment
		def render_interactive_table():
			def register_download():
				page_no = int(st.session_state.get('current_page', 1) or 1)
				page_downloads = st.session_state.get('page_downloads', {})
				if page_no in page_downloads:
					page_downloads[page_no]['download_completed'] = True
					st.session_state['page_downloads'] = page_downloads
				st.session_state['download_completed'] = True

			total_rows = len(search_results)
			page_size = int(st.session_state.get('page_size', 10) or 10)
			total_pages = max(1, (total_rows + page_size - 1) // page_size)
			current_page = int(st.session_state.get('current_page', 1) or 1)
			if current_page < 1:
				current_page = 1
			if current_page > total_pages:
				current_page = total_pages
			st.session_state['current_page'] = current_page

			page_start = (current_page - 1) * page_size
			page_end = min(page_start + page_size, total_rows)
			page_indices = list(range(page_start, page_end))
			st.session_state['current_page_indices'] = page_indices
			st.session_state['select_all_main'] = bool(page_indices) and all(st.session_state.get(f'chk_{i}', False) for i in page_indices)
			st.session_state['select_all_state'] = st.session_state['select_all_main']

			st.markdown('<div class="table-row-container">', unsafe_allow_html=True)

			header_cols = st.columns([0.05, 0.08, 0.32, 0.1, 0.18, 0.34])
			header_style = 'background-color: #f0f0f0; padding: 6px 12px; font-weight: bold; text-align: center; border-radius: 4px;'

			with header_cols[0]:
				st.checkbox('', key='select_all_main', on_change=toggle_select_all)
			with header_cols[1]:
				st.markdown(f'<div style="{header_style}">No</div>', unsafe_allow_html=True)
			with header_cols[2]:
				st.markdown(f'<div style="{header_style}">업체명</div>', unsafe_allow_html=True)
			with header_cols[3]:
				st.markdown(f'<div style="{header_style}">지역</div>', unsafe_allow_html=True)
			with header_cols[4]:
				st.markdown(f'<div style="{header_style}">상태</div>', unsafe_allow_html=True)
			with header_cols[5]:
				st.markdown(f'<div style="{header_style}">파일명</div>', unsafe_allow_html=True)

			st.markdown('<div style="margin-bottom: 12px;"></div>', unsafe_allow_html=True)

			status_text_placeholders = {}
			status_bar_placeholders = {}
			filename_placeholders = {}

			def _render_row_status_if_visible(row_idx):
				text_ph = status_text_placeholders.get(row_idx)
				bar_ph = status_bar_placeholders.get(row_idx)
				if text_ph is not None and bar_ph is not None:
					_render_status_cell(row_idx, text_ph, bar_ph)

			def _write_row_filename_if_visible(row_idx, value):
				name_ph = filename_placeholders.get(row_idx)
				if name_ph is not None:
					name_ph.write(value)

			for i in page_indices:
				result = search_results[i]
				if f'chk_{i}' not in st.session_state:
					st.session_state[f'chk_{i}'] = False
				if f'status_{i}' not in st.session_state:
					st.session_state[f'status_{i}'] = ''
				if f'progress_{i}' not in st.session_state:
					st.session_state[f'progress_{i}'] = 0
				if f'filename_{i}' not in st.session_state:
					st.session_state[f'filename_{i}'] = ''

				cols = st.columns([0.05, 0.08, 0.32, 0.1, 0.18, 0.34])

				with cols[0]:
					st.checkbox('', key=f'chk_{i}', label_visibility='collapsed', on_change=check_individual_toggle)
				with cols[1]:
					st.markdown(f'<div style="text-align: center;">{i + 1}</div>', unsafe_allow_html=True)
				with cols[2]:
					st.write(result.get('company_name', ''))
				with cols[3]:
					st.markdown(f'<div style="text-align: center;">{result.get("region", "") or ""}</div>', unsafe_allow_html=True)
				with cols[4]:
					status_text_placeholders[i] = st.empty()
					status_bar_placeholders[i] = st.empty()
					_render_status_cell(i, status_text_placeholders[i], status_bar_placeholders[i])
				with cols[5]:
					filename_placeholders[i] = st.empty()
					filename_placeholders[i].write(st.session_state[f'filename_{i}'])

			# 하단 페이징 컨트롤 (이미지 스타일: « ‹ 1 2 › »)
			window_size = 5
			half_window = window_size // 2
			start_page = max(1, current_page - half_window)
			end_page = min(total_pages, start_page + window_size - 1)
			start_page = max(1, end_page - window_size + 1)
			visible_pages = list(range(start_page, end_page + 1))

			st.markdown('<div class="page-nav-anchor"></div>', unsafe_allow_html=True)
			outer_left, outer_center, outer_right = st.columns([0.33, 0.34, 0.33])
			with outer_center:
				is_generating = st.session_state.get('file_generation_started', False)
				control_count = len(visible_pages) + 4
				control_cols = st.columns([1] * control_count)
				control_idx = 0

				with control_cols[control_idx]:
					if st.button('«', key='btn_page_first', use_container_width=True, type='tertiary', disabled=is_generating):
						if current_page > 1:
							st.session_state['current_page'] = 1
							st.rerun()
				control_idx += 1

				with control_cols[control_idx]:
					if st.button('‹', key='btn_page_prev', use_container_width=True, type='tertiary', disabled=is_generating):
						if current_page > 1:
							st.session_state['current_page'] = current_page - 1
							st.rerun()
				control_idx += 1

				for page_no in visible_pages:
					with control_cols[control_idx]:
						if page_no == current_page:
							st.button(str(page_no), key=f'btn_page_{page_no}', use_container_width=True, type='tertiary', disabled=True)
						else:
							if st.button(str(page_no), key=f'btn_page_{page_no}', use_container_width=True, type='tertiary', disabled=is_generating):
								st.session_state['current_page'] = page_no
								st.rerun()
					control_idx += 1

				with control_cols[control_idx]:
					if st.button('›', key='btn_page_next', use_container_width=True, type='tertiary', disabled=is_generating):
						if current_page < total_pages:
							st.session_state['current_page'] = current_page + 1
							st.rerun()
				control_idx += 1

				with control_cols[control_idx]:
					if st.button('»', key='btn_page_last', use_container_width=True, type='tertiary', disabled=is_generating):
						if current_page < total_pages:
							st.session_state['current_page'] = total_pages
							st.rerun()

			st.markdown('</div>', unsafe_allow_html=True)

			st.markdown('<div class="bottom-card-anchor"></div>', unsafe_allow_html=True)
			spacer_col, col_btn_gen, col_btn_down = st.columns([0.7, 0.15, 0.15], vertical_alignment='center')

			with col_btn_gen:
				is_generating = st.session_state.get('file_generation_started', False)
				gen_clicked_bottom = st.button('📝 파일 생성', key='btn_gen', use_container_width=True, disabled=is_generating)

			with col_btn_down:
				page_download = st.session_state.get('page_downloads', {}).get(current_page)
				if page_download and page_download.get('data') is not None:
					st.download_button(
						label='📥 다운로드',
						data=page_download.get('data'),
						file_name=page_download.get('filename', ''),
						mime=page_download.get('mime', 'application/zip'),
						key='btn_down',
						use_container_width=True,
						on_click=register_download,
					)
				else:
					st.button('📥 다운로드', key='btn_down_disabled', use_container_width=True, disabled=True)

			if gen_clicked_bottom:
				selected_idxs = [idx for idx in page_indices if st.session_state.get(f'chk_{idx}', False)]

				if selected_idxs:
					st.session_state['file_generation_started'] = True
					st.session_state['active_generation_indices'] = list(selected_idxs)
					st.session_state['file_generation_completed'] = False
					page_downloads = st.session_state.get('page_downloads', {})
					page_downloads[current_page] = {
						'data': None,
						'filename': '',
						'mime': 'application/zip',
						'target_indices': list(selected_idxs),
						'download_completed': False,
					}
					st.session_state['page_downloads'] = page_downloads
					st.session_state['download_completed'] = False

					for i in selected_idxs:
						st.session_state[f'status_{i}'] = ''
						st.session_state[f'progress_{i}'] = 0
						st.session_state[f'filename_{i}'] = ''
						_render_row_status_if_visible(i)
						_write_row_filename_if_visible(i, '')

					for idx in selected_idxs:
						item_data = search_results[idx]
						keyword = item_data.get('search_keyword', '')
						bplc_id = item_data.get('bplcId', '')
						company_name = item_data.get('company_name', '')
						region = item_data.get('region', '')
						company_name_key = company_name
						global_same_name_count = sum(
							1
							for row in search_results
							if row.get('company_name', '') == company_name_key
						)
						use_region = global_same_name_count > 1 or item_data.get('use_region_in_filename', False)

						if bplc_id and '검색 결과 없음' not in company_name:
							try:
								st.session_state[f'status_{idx}'] = '처리중'
								st.session_state[f'progress_{idx}'] = 5
								_render_row_status_if_visible(idx)

								def _company_progress(completed, total, row_idx=idx):
									if total <= 0:
										pct = 90
									else:
										pct = min(90, max(5, int((completed / total) * 90)))
									st.session_state[f'progress_{row_idx}'] = pct
									st.session_state[f'status_{row_idx}'] = '처리중'
									_render_row_status_if_visible(row_idx)

								session = requests.Session()
								fetch_search_page(session)
								company_name_full, final_bytes = create_final_excel_bytes_for_company(
									session=session,
									keyword=keyword,
									company_item={'bplcId': bplc_id, 'bplcNm': company_name},
									progress_callback=_company_progress,
								)

								if use_region and region and region != '':
									filename = f"{sanitize_filename(company_name_full)}_{sanitize_filename(region)}.xlsx"
								else:
									filename = f"{sanitize_filename(company_name_full)}.xlsx"

								if 'files' not in st.session_state:
									st.session_state['files'] = {}
								st.session_state['files'][filename] = final_bytes

								st.session_state[f'status_{idx}'] = '✅'
								st.session_state[f'progress_{idx}'] = 100
								st.session_state[f'filename_{idx}'] = filename
								_render_row_status_if_visible(idx)
								_write_row_filename_if_visible(idx, filename)
							except Exception as exc:
								error_msg = str(exc)
								st.session_state[f'status_{idx}'] = '❌'
								st.session_state[f'progress_{idx}'] = max(int(st.session_state.get(f'progress_{idx}', 0) or 0), 1)
								st.session_state[f'filename_{idx}'] = f'{company_name}_ERROR ({error_msg})'
								_render_row_status_if_visible(idx)
								_write_row_filename_if_visible(idx, f'{company_name}_ERROR ({error_msg})')

						time.sleep(0.1)

					selected_filenames = []
					for idx in selected_idxs:
						filename = st.session_state.get(f'filename_{idx}', '')
						if filename and '_ERROR' not in filename:
							selected_filenames.append(filename)

					if selected_filenames:
						if len(selected_filenames) == 1:
							single_filename = selected_filenames[0]
							file_bytes = st.session_state['files'].get(single_filename)
							page_downloads = st.session_state.get('page_downloads', {})
							page_downloads[current_page] = {
								'data': file_bytes,
								'filename': single_filename,
								'mime': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
								'target_indices': list(selected_idxs),
								'download_completed': False,
							}
							st.session_state['page_downloads'] = page_downloads
						else:
							zip_bytes = zip_files(st.session_state['files'], selected_filenames)
							now = datetime.now().strftime('%Y%m%d_%H%M%S')
							keywords_str = st.session_state['keywords'].replace(', ', '_').replace(' ', '_')
							page_downloads = st.session_state.get('page_downloads', {})
							page_downloads[current_page] = {
								'data': zip_bytes,
								'filename': f'{keywords_str}_{now}.zip',
								'mime': 'application/zip',
								'target_indices': list(selected_idxs),
								'download_completed': False,
							}
							st.session_state['page_downloads'] = page_downloads

						st.session_state['file_generation_completed'] = True
						st.session_state['file_generation_started'] = False
						st.session_state['active_generation_indices'] = []
						st.rerun()
					else:
						st.session_state['file_generation_started'] = False
						st.session_state['active_generation_indices'] = []

		render_interactive_table()


if __name__ == '__main__':
	main_ui()
